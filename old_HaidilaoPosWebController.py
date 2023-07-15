import asyncio
import time
import requests
import json
import os
from pathlib import Path
import shutil
from openpyxl import Workbook, load_workbook
import datetime


def daily_work():
    current_requesting_date = input("enter the date you currently working on in format of yyyy-mm-dd: ")
    split_times = current_requesting_date.split("-")
    next_day_date = str(datetime.datetime(int(split_times[0]), int(split_times[1]), int(split_times[2])) + datetime.timedelta(days=1)).split(" ")[0]
    # create new daily directory
    print("create new daily directory")
    current_dir = f"daily_data_folder/{current_requesting_date}"
    if Path(current_dir).exists():
        try:
            shutil.rmtree(current_dir)
        except Exception as e:
            shutil.rmtree(current_dir)
        time.sleep(2)
    os.mkdir(current_dir)
    current_dir = f"daily_data_folder/{current_requesting_date}/"
    s = requests.Session()

    # login, and get my personal info
    haidilao_pos_user_name = "10312229"
    haidilao_pos_password = "6/9USW7DQhgs+h1inlAmXRX+q9oVt2BkL9qp1LtG5og="
    print("login, and get my personal info")
    payload = {
        "userName": haidilao_pos_user_name,
        "password": haidilao_pos_password,
        "osId": 0,
        "type": "web",
        # "sig": "7b909d5d8e05f146bb89f81fc661c481"
    }
    pos_res_info = s.post("https://pos.haidilao.com:8036/login", json=payload)  # this is raw info
    json_pos_res_info = json.loads(pos_res_info.content)  # this is result converted raw info to json info
    s.headers.update({"token": json.loads(pos_res_info.content)["obj"]["Token"]})  # add temperary token into session header

    # daily data close
    print("daily data close")
    params = {
        "shopId": json_pos_res_info["obj"]["shopList"][0]["id"],
        "creatorId": json_pos_res_info["obj"]["empId"],
        "companyId": json_pos_res_info["obj"]["companyInfo"]["id"],
        "date": current_requesting_date,
        "userName": json_pos_res_info["obj"]["userName"],
        "userId": json_pos_res_info["obj"]["userId"],
        # "sig": "82f3b3fd34938c8329baadffbb44f526"
    }
    daily_close_url = "https://pos.haidilao.com:8032/clearSummary/triggerClearSummary"
    data ={}
    daily_close_result = s.get(daily_close_url, params=params)

    # get daily close info
    print("get daily close info")
    params = {
        "shopId": json_pos_res_info["obj"]["shopList"][0]["id"],
        "creatorId": json_pos_res_info["obj"]["empId"],
        "date": current_requesting_date,
        "userName": json_pos_res_info["obj"]["userName"],
        "userId": json_pos_res_info["obj"]["userId"],
        # "sig": "6426f7fac9480dd6b10002771c91336e"
    }
    daily_close_info_url = "https://pos.haidilao.com:8032/clearSummary/getClearSummaryReport"
    daily_close_info_result = s.get(daily_close_info_url, params=params)
    json_daily_close_info_result = json.loads(daily_close_info_result.content)

    # fill updated daily close info
    print("fill updated daily close info")
    big_note = 0
    account_summary_list = json_daily_close_info_result["data"]["accountSummaryList"]
    for item in account_summary_list:
        if item["payType"] == "cash":
            big_note = item["payMoney"]
            item["receiveMoney"] = item["payMoney"]
        else:
            item["receiveMoney"] = item["payMoney"]
    payload = {
        "accountSummaryList": account_summary_list,
        "accountDate": json_daily_close_info_result["data"]["accountDate"],
        "creatorId": json_pos_res_info["obj"]["empId"],
        "companyId": json_pos_res_info["obj"]["companyInfo"]["id"],
        "createTime": json_daily_close_info_result["data"]["createTime"],
        "id": json_daily_close_info_result["data"]["id"],
        "userName": json_pos_res_info["obj"]["userName"],
        "userId": json_pos_res_info["obj"]["userId"],
        "shopId": json_pos_res_info["obj"]["shopList"][0]["id"],
        # "sig": "64a29835bc1b1b221afc14685275dd77",
        "receiverName": "Hongming Wang",
        "receiverPhone": "7788611008",
        "bigNote": big_note,
        "clearerName": "Hongming Wang",
        "clearerPhone": "7788611008",
        "receiveDifference": 0,
        "receiveSummary": json_daily_close_info_result["data"]["accountSummary"],
        "savingMoney": big_note
    }
    daily_fill_url = "https://pos.haidilao.com:8032/clearSummary/addClearSummaryReport"
    daily_fill_url_result = s.post(daily_fill_url, json=payload)

    # download daily close book xlsx
    print("download daily close book xlsx")
    params = {
        "shopId": json_pos_res_info["obj"]["shopList"][0]["id"],
        "shopName": json_pos_res_info["obj"]["shopList"][0]["shopName"],
        "creatorId": json_pos_res_info["obj"]["empId"],
        "date": current_requesting_date,
        "userName": json_pos_res_info["obj"]["userName"],
        "userId": json_pos_res_info["obj"]["userId"],
        "sig": "cccd9178afd3799154bb069d04dcbac2"
    }
    download_daily_close_book_url = "https://pos.haidilao.com:8032/clearSummary/clearSummaryExportExcel"
    download_daily_close_book_result = s.post(download_daily_close_book_url, params=params)
    json_download_daily_close_book_result = json.loads(download_daily_close_book_result.content)

    download_xlsx_url = json_download_daily_close_book_result["data"]["url"]
    xlsx_file = s.get(download_xlsx_url)
    filename = f"{current_dir}{current_requesting_date} canada1 daily close data.xlsx"

    with open(filename, "wb") as f:
        for chunk in xlsx_file.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)
    f.close()

    # create new daily close book
    print("create new daily close book")
    daily_close_book_template_path = "templates/Close_Book.xlsx"
    daily_close_book_path = f"{current_dir}{current_requesting_date}canada1 daily close book.xlsx"
    shutil.copy(daily_close_book_template_path, daily_close_book_path)

    # import daily close datas to daily close book
    print("import daily close datas to daily close book")
    daily_close_book_workbook = load_workbook(daily_close_book_path)
    daily_close_book_workbook_ws = daily_close_book_workbook.active
    daily_close_data_workbook = load_workbook(filename)
    daily_close_data_workbook_ws = daily_close_data_workbook.active

    for col in range(1, 35):
        if daily_close_data_workbook_ws["D"+str(col)].value == "现金":
            daily_close_book_workbook_ws["G7"] = float(daily_close_data_workbook_ws["F"+str(col)].value)
        elif daily_close_data_workbook_ws["D"+str(col)].value == "信用卡":
            daily_close_book_workbook_ws["G8"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
        elif daily_close_data_workbook_ws["D"+str(col)].value == "饭团":
            daily_close_book_workbook_ws["G10"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
        elif daily_close_data_workbook_ws["D"+str(col)].value == "线下支付宝":
            if daily_close_book_workbook_ws["G9"].value is not None:
                daily_close_book_workbook_ws["G9"] = float(daily_close_data_workbook_ws["F" + str(col)].value) + \
                                                     float(daily_close_book_workbook_ws["G9"].value)
            else:
                daily_close_book_workbook_ws["G9"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
        elif daily_close_data_workbook_ws["D"+str(col)].value == "线下微信":
            if daily_close_book_workbook_ws["G9"].value is not None:
                daily_close_book_workbook_ws["G9"] = float(daily_close_data_workbook_ws["F" + str(col)].value) + \
                                                     float(daily_close_book_workbook_ws["G9"].value)
            else:
                daily_close_book_workbook_ws["G9"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
        elif daily_close_data_workbook_ws["D"+str(col)].value == "Snappy":
            daily_close_book_workbook_ws["G16"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
        elif daily_close_data_workbook_ws["D"+str(col)].value == "uber":
            daily_close_book_workbook_ws["G14"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
        elif daily_close_data_workbook_ws["D"+str(col)].value == "Vanpeople":
            daily_close_book_workbook_ws["G19"] = float(daily_close_data_workbook_ws["F" + str(col)].value)

    # import daily summary data to daily close book
    params = {
        "groupCollect1": 1,
        "date": f"{current_requesting_date},{current_requesting_date}",
        "pageSize": 30,
        "pageNum": 1,
        "shopId": json_pos_res_info["obj"]["shopList"][0]["id"],
        "userName": json_pos_res_info["obj"]["userName"],
        "userId": json_pos_res_info["obj"]["userId"],
    }
    daily_summary_data_url = "https://pos.haidilao.com:8032/repBusinessTable/list"
    daily_summary_data_result = s.get(daily_summary_data_url, params=params)
    json_daily_summary_data_result = json.loads(daily_summary_data_result.content)

    daily_close_book_workbook_ws["B2"] = float(json_daily_summary_data_result["data"]["list"][0]["payShop"])
    daily_close_book_workbook_ws["B3"] = float(json_daily_summary_data_result["data"]["list"][0]["paySend"])
    daily_close_book_workbook_ws["B1"] = float(json_daily_summary_data_result["data"]["list"][0]["payOut"])

    # import third party data to daily close book
    # import Uber Eats Data
    get_uber_data()

    # import Fantuan Data
    get_fantuan_data()

    daily_close_book_workbook.save(daily_close_book_path)

    # print(daily_close_result.content)
    # print(daily_close_info_result.content)
    # print(s.headers)
    # print(json_pos_res_info)


def get_uber_data():
    return
    current_requesting_date = "2022-02-10"
    split_times = current_requesting_date.split("-")
    next_day_date = str(datetime.datetime(int(split_times[0]), int(split_times[1]), int(split_times[2]))+datetime.timedelta(days=1)).split(" ")[0]
    s = requests.Session()
    payload = {
        "report_type": "PAYMENT_DETAILS_REPORT",
        "store_uuids": ["6caef017-314a-453a-b658-e5d6957532a7"],
        "group_uuids": ["6caef017-314a-453a-b658-e5d6957532a7"],
        "start_date": current_requesting_date,
        "end_date": next_day_date
    }
    uber_request_data_url = "https://api.uber.com/v1/eats/report"
    uber_data_result = s.post(uber_request_data_url, json=payload)
    # json_uber_data_result = json.loads(uber_data_result.content)
    print(uber_data_result.content)


def get_fantuan_data():
    current_requesting_date = "2022-02-10"
    split_times = current_requesting_date.split("-")
    next_day_date = str(datetime.datetime(int(split_times[0]), int(split_times[1]), int(split_times[2])) + datetime.timedelta(
        days=1)).split(" ")[0]
    s = requests.Session()
    print(f"{datetime.datetime.timestamp(datetime.datetime.now())}")
    s.headers.update({
        "sign": "b52f0cd6a88ab2dea7657f235472bf4f200ea2fa01200aab698291052307b0ab",
        "meta": '{"ucId": "", "restaurantId": "", "deviceId": "19be44c7-add4-48b8-9e34-901f47944c67", "sn": ""}',
        "nonce": "e5629cddc5ea4eb28e71dafb5ed64722",
        "timestamp": f"{datetime.datetime.timestamp(datetime.datetime.now())}"
    })
    payload = {
        "nickname": "6043706665",
        "password": "Qz059522778157",
        "country": "CA",
        "loginMethod": "name_pwd",
        "encrypt": False
        }
    fantuan_request_data_url = "https://ca-gateway.fantuan.ca/api/usercenter/public/login"
    fantuan_data_result = s.post(fantuan_request_data_url, json=payload)
    json_fantuan_data_result = json.loads(fantuan_data_result.content)
    print(json_fantuan_data_result)


daily_work()