import openpyxl
import json

data_sheet_path = 'website/data/employee_data.json'
import_from_path = 'target_file.xlsx'

employees = {}


class Employee:
    employee_number = ""
    legal_name = ""
    prefer_name = ""
    department = ""
    server_tips_ratio = 0
    dish_prepare_tips_ratio = 0
    dish_runner_tips_ratio = 0
    appetizer_tips_ratio = 0
    cleaner_tips_ratio = 0
    host_tips_ratios = 0
    noodle_dance_tips_ratios = 0
    dish_prepare_work_count_ratio = 0
    dish_runner_work_count_ratio = 0
    appetizer_work_count_ratio = 0


wb = openpyxl.load_workbook(import_from_path)
ws = wb.active

for row in range(1, ws.max_row):
    if ws[f"B{row}"].value is not None and str(ws[f"B{row}"].value).isdigit():
        employees[str(ws[f"B{row}"].value)] = Employee()
        current_employee = employees[str(ws[f"B{row}"].value)]
        current_employee.employee_number = str(ws[f"B{row}"].value)
        current_employee.legal_name = str(ws[f"D{row}"].value)
        current_employee.prefer_name = str(ws[f"C{row}"].value)
        current_employee.server_tips_ratio = str(ws[f"E{row}"].value)
        current_employee.dish_prepare_tips_ratio = str(ws[f"F{row}"].value)
        current_employee.dish_runner_tips_ratio = str(ws[f"G{row}"].value)
        current_employee.appetizer_tips_ratio = str(ws[f"H{row}"].value)
        current_employee.cleaner_tips_ratio = str(ws[f"I{row}"].value)
        current_employee.host_tips_ratios = str(ws[f"J{row}"].value)
        current_employee.noodle_dance_tips_ratios = str(ws[f"K{row}"].value)
        current_employee.dish_prepare_work_count_ratio = str(ws[f"F{row}"].value)
        current_employee.dish_runner_work_count_ratio = str(ws[f"G{row}"].value)
        current_employee.appetizer_work_count_ratio = str(ws[f"H{row}"].value)
        current_employee = json.dumps(current_employee.__dict__)
        employees[str(ws[f"B{row}"].value)] = current_employee
wb.close()


json_string = json.dumps(employees)
target_data_file = open(data_sheet_path, "w")
target_data_file.truncate()
target_data_file.write(json_string)


