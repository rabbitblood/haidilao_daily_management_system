import json
import os
import shutil
import time

import openpyxl
import requests

from website import data_handler
import datetime

file_path = os.path.dirname(__file__)
destination_path = data_handler.get_current_daily_file_path()


def init_files():
    global destination_path
    destination_path = data_handler.get_current_daily_file_path()
    print("创建文件夹")
    for file in os.listdir(file_path + "/temp_file"):
        os.remove(file_path + "/temp_file/" + file)
    try:
        shutil.rmtree(destination_path)
    except:
        pass
    time.sleep(1)
    print("拷贝模板文档")
    shutil.copytree((file_path + "/static/file_templates"), destination_path)
    for files in os.listdir(destination_path):
        os.rename(destination_path + "/" + files, destination_path + "/" + data_handler.get_server_date() + files)
    print("清除日结表表格内之前的数据")
    copied_close_book_path = None  # 收账本模板
    for file in os.listdir(destination_path + f"/{data_handler.get_server_date()}收账明细"):
        if "收账本" in file:
            copied_close_book_path = destination_path + f"/{data_handler.get_server_date()}收账明细/" + file
    copied_close_book_wb = openpyxl.load_workbook(copied_close_book_path)
    copied_daily_close_book_workbook_ws = copied_close_book_wb.active
    for row in copied_daily_close_book_workbook_ws['B1':'B3']:
        for cell in row:
            cell.value = None
    for row in copied_daily_close_book_workbook_ws['C7':'G21']:
        for cell in row:
            cell.value = None
    for row in copied_daily_close_book_workbook_ws['F3':'F4']:
        for cell in row:
            cell.value = None
    for row in copied_daily_close_book_workbook_ws['L3':'O10']:
        for cell in row:
            cell.value = None
    for row in copied_daily_close_book_workbook_ws['R1':'R5']:
        for cell in row:
            cell.value = None
    for row in copied_daily_close_book_workbook_ws["I7":"J21"]:
        for cell in row:
            cell.value = None
    copied_close_book_wb.save(copied_close_book_path)
    copied_close_book_wb.close()

    # 厨房计件文件
    kitchen_work_count_xlsx_file_path = ""
    for file in os.listdir(destination_path):
        if "厨房计件" in file:
            kitchen_work_count_xlsx_file_path = destination_path + "/" + file
            break
    time.sleep(1)
    kitchen_work_count_xlsx_wb = openpyxl.load_workbook(kitchen_work_count_xlsx_file_path)
    kitchen_work_count_xlsx_ws = kitchen_work_count_xlsx_wb.active

    for row in kitchen_work_count_xlsx_ws['B1':'B2']:
        for cell in row:
            cell.value = None
    for row in kitchen_work_count_xlsx_ws['A6':'C9']:
        for cell in row:
            cell.value = None
    for row in kitchen_work_count_xlsx_ws['A14':'C17']:
        for cell in row:
            cell.value = None
    for row in kitchen_work_count_xlsx_ws['A22':'C25']:
        for cell in row:
            cell.value = None
    for row in kitchen_work_count_xlsx_ws['A30':'C33']:
        for cell in row:
            cell.value = None
    for row in kitchen_work_count_xlsx_ws['A38':'C41']:
        for cell in row:
            cell.value = None
    for row in kitchen_work_count_xlsx_ws['A46':'C53']:
        for cell in row:
            cell.value = None

    kitchen_work_count_xlsx_wb.save(kitchen_work_count_xlsx_file_path)
    kitchen_work_count_xlsx_wb.close()

    # 传配计件文件
    busser_work_count_xlsx_file_path = ""
    for file in os.listdir(destination_path):
        if "传配组打卡计件" in file:
            busser_work_count_xlsx_file_path = destination_path + "/" + file
            break
    time.sleep(1)
    busser_work_count_xlsx_wb = openpyxl.load_workbook(busser_work_count_xlsx_file_path)
    busser_work_count_xlsx_ws = busser_work_count_xlsx_wb.active

    for row in busser_work_count_xlsx_ws['A13':'E49']:
        for cell in row:
            cell.value = None
    for row in busser_work_count_xlsx_ws['B2':'B7']:
        for cell in row:
            cell.value = None

    busser_work_count_xlsx_wb.save(busser_work_count_xlsx_file_path)
    busser_work_count_xlsx_wb.close()

    # 小费文件
    tips_count_xlsx_file_path = ""
    for file in os.listdir(destination_path):
        if "小费计算表格" in file:
            tips_count_xlsx_file_path = destination_path + "/" + file
            break
    time.sleep(1)
    tips_count_xlsx_wb = openpyxl.load_workbook(tips_count_xlsx_file_path)
    tips_count_xlsx_ws = tips_count_xlsx_wb.active

    tips_count_xlsx_wb.save(tips_count_xlsx_file_path)
    tips_count_xlsx_wb.close()

    # init close data dict
    daily_data = data_handler.get_daily_data()
    daily_data["close_data"] = {}
    data_handler.save_daily_data_as_json(json.dumps(daily_data))


def busser_work_count():
    # daily
    employee_data = data_handler.get_employee_data()
    daily_data = data_handler.get_daily_data()
    busser_work_count_xlsx_file_path = ""
    daily_busser_work_count = {}
    for file in os.listdir(destination_path):
        if "传配组打卡计件" in file:
            busser_work_count_xlsx_file_path = destination_path + "/" + file
            break
    time.sleep(1)

    busser_work_count_xlsx_wb = openpyxl.load_workbook(busser_work_count_xlsx_file_path)
    busser_work_count_xlsx_ws = busser_work_count_xlsx_wb.active

    print(daily_data)

    # 从daily data得到general data
    morning_table_count = int(daily_data["general_data"]["morning_table_count"])
    evening_table_count = int(daily_data["general_data"]["evening_table_count"])
    take_out_count = int(daily_data["general_data"]["take_out_count"])
    morning_busser_help_count = float(daily_data["general_data"]["morning_busser_help_count"])
    evening_busser_help_count = float(daily_data["general_data"]["evening_busser_help_count"])

    # 计算出早晚班总计件数量
    morning_total_money = ((morning_table_count - (take_out_count / 2) + (
            take_out_count / 3 / 2)) * 7) - morning_busser_help_count
    evening_total_money = ((evening_table_count - (take_out_count / 2) + (
            take_out_count / 3 / 2)) * 7) - evening_busser_help_count

    daily_busser_employee_data = {}

    # 填写excel表格内的日信息
    busser_work_count_xlsx_ws['B2'] = data_handler.get_server_date()
    busser_work_count_xlsx_ws['B3'] = morning_table_count
    busser_work_count_xlsx_ws['B4'] = evening_table_count
    busser_work_count_xlsx_ws['B5'] = take_out_count
    busser_work_count_xlsx_ws['B6'] = morning_busser_help_count
    busser_work_count_xlsx_ws['B7'] = evening_busser_help_count

    start_row = 13
    for person in daily_data["dish_runner_work_hour"]:
        employee_number = str(person).split(" ")[0]
        employee_name = str(person).split(" ", 1)[1]
        employee_busser_work_count_ratio = float(
            employee_data[str(person).split(" ")[0]]["dish_runner_work_count_ratio"])
        employee_morning_work_hour = float(daily_data["dish_runner_work_hour"][person]["morning_worked_hour"])
        employee_evening_worked_hour = float(daily_data["dish_runner_work_hour"][person]["evening_worked_hour"])

        # 在每日计件员工内加入此员工 用来计算每日员工计件所得
        daily_busser_employee_data[employee_number] = {"employee_number": employee_number,
                                                       "employee_busser_work_count_ratio": employee_busser_work_count_ratio,
                                                       "employee_morning_work_hour": employee_morning_work_hour,
                                                       "employee_evening_worked_hour": employee_evening_worked_hour}

        busser_work_count_xlsx_ws[f"A{start_row}"] = employee_number
        busser_work_count_xlsx_ws[f"B{start_row}"] = employee_name
        try:
            busser_work_count_xlsx_ws[f"C{start_row}"] = employee_busser_work_count_ratio
        except Exception as e:
            busser_work_count_xlsx_ws[f"C{start_row}"] = 0
        busser_work_count_xlsx_ws[f"D{start_row}"] = employee_morning_work_hour
        busser_work_count_xlsx_ws[f"E{start_row}"] = employee_evening_worked_hour
        start_row += 1

    # 计算每日员工计件所得
    total_morning_busser_ratio = 0
    total_evening_busser_ratio = 0

    for employee in daily_busser_employee_data:
        total_morning_busser_ratio += \
            (float(daily_busser_employee_data[employee]["employee_busser_work_count_ratio"]) *
             float(daily_busser_employee_data[employee]["employee_morning_work_hour"]))
        total_evening_busser_ratio += \
            (float(daily_busser_employee_data[employee]["employee_busser_work_count_ratio"]) *
             float(daily_busser_employee_data[employee]["employee_evening_worked_hour"]))

    for employee in daily_busser_employee_data:
        current_employee_morning_work_count = float(
            morning_total_money / total_morning_busser_ratio * daily_busser_employee_data[employee][
                "employee_busser_work_count_ratio"] * daily_busser_employee_data[employee][
                "employee_morning_work_hour"])

        current_employee_evening_work_count = float(
            evening_total_money / total_evening_busser_ratio * daily_busser_employee_data[employee][
                "employee_busser_work_count_ratio"] * daily_busser_employee_data[employee][
                "employee_evening_worked_hour"])

        daily_busser_work_count[daily_busser_employee_data[employee]["employee_number"]] = round(
            current_employee_morning_work_count + current_employee_evening_work_count, 2)

    # 结束计算，关闭文件
    busser_work_count_xlsx_wb.save(busser_work_count_xlsx_file_path)
    busser_work_count_xlsx_wb.close()
    daily_data["close_data"]["daily_busser_work_count"] = daily_busser_work_count
    data_handler.save_daily_data_as_json(json.dumps(daily_data))


def kitchen_work_count():
    employee_data = data_handler.get_employee_data()
    daily_data = data_handler.get_daily_data()
    daily_kitchen_work_count = {}

    kitchen_work_count_xlsx_file_path = ""
    for file in os.listdir(destination_path):
        if "厨房计件" in file:
            kitchen_work_count_xlsx_file_path = destination_path + "/" + file
            break
    time.sleep(1)
    kitchen_work_count_xlsx_wb = openpyxl.load_workbook(kitchen_work_count_xlsx_file_path)
    kitchen_work_count_xlsx_ws = kitchen_work_count_xlsx_wb.active

    # 日期 桌数
    kitchen_work_count_xlsx_ws["B2"] = data_handler.get_server_date()
    day_table_count = int(daily_data["general_data"]["evening_table_count"]) + int(
        daily_data["general_data"]["morning_table_count"])
    kitchen_work_count_xlsx_ws["B1"] = day_table_count

    # 上菜口
    row_number = 6
    dish_prepare_window_work_count_per_table = 1.35
    dish_prepare_window_window_total_hour = 0
    dish_prepare_window_total_amount = day_table_count * dish_prepare_window_work_count_per_table
    for number in range(1, 5):
        if daily_data["dish_prepare_work_hour"][f"dish_prepare_window{number}"] != "":
            current_employee_number = \
                str(daily_data["dish_prepare_work_hour"][f"dish_prepare_window{number}"]).split(" ")[0]
            current_employee_name = \
                str(daily_data["dish_prepare_work_hour"][f"dish_prepare_window{number}"]).split(" ", 1)[1]
            current_employee_work_hour_today = float(
                daily_data["dish_prepare_work_hour"][f"dish_prepare_window{number}_work_hour"])
            kitchen_work_count_xlsx_ws[f"A{row_number}"] = current_employee_number
            kitchen_work_count_xlsx_ws[f"B{row_number}"] = current_employee_name
            kitchen_work_count_xlsx_ws[f"C{row_number}"] = current_employee_work_hour_today
            dish_prepare_window_window_total_hour += current_employee_work_hour_today
            row_number += 1
    for number in range(1, 5):
        if daily_data["dish_prepare_work_hour"][f"dish_prepare_window{number}"] != "":
            current_employee_number = \
                str(daily_data["dish_prepare_work_hour"][f"dish_prepare_window{number}"]).split(" ")[0]
            current_employee_work_hour_today = float(
                daily_data["dish_prepare_work_hour"][f"dish_prepare_window{number}_work_hour"])
            if not daily_kitchen_work_count.get(current_employee_number):
                daily_kitchen_work_count[current_employee_number] = 0
            daily_kitchen_work_count[current_employee_number] += round(
                dish_prepare_window_total_amount / dish_prepare_window_window_total_hour * current_employee_work_hour_today,
                2)

    # 羊肉间
    row_number = 14
    lamb_room_work_count_per_table = 1.25
    lamb_room_window_total_hour = 0
    lamb_room_total_amount = day_table_count * lamb_room_work_count_per_table
    for number in range(1, 5):
        if daily_data["dish_prepare_work_hour"][f"lamb_room{number}"] != "":
            current_employee_number = str(daily_data["dish_prepare_work_hour"][f"lamb_room{number}"]).split(" ")[0]
            current_employee_name = str(daily_data["dish_prepare_work_hour"][f"lamb_room{number}"]).split(" ", 1)[1]
            current_employee_work_hour_today = float(
                daily_data["dish_prepare_work_hour"][f"lamb_room{number}_work_hour"])
            kitchen_work_count_xlsx_ws[f"A{row_number}"] = current_employee_number
            kitchen_work_count_xlsx_ws[f"B{row_number}"] = current_employee_name
            kitchen_work_count_xlsx_ws[f"C{row_number}"] = current_employee_work_hour_today
            lamb_room_window_total_hour += current_employee_work_hour_today
            row_number += 1
    for number in range(1, 5):
        if daily_data["dish_prepare_work_hour"][f"lamb_room{number}"] != "":
            current_employee_number = str(daily_data["dish_prepare_work_hour"][f"lamb_room{number}"]).split(" ")[0]
            current_employee_work_hour_today = float(
                daily_data["dish_prepare_work_hour"][f"lamb_room{number}_work_hour"])
            if not daily_kitchen_work_count.get(current_employee_number):
                daily_kitchen_work_count[current_employee_number] = 0
            daily_kitchen_work_count[current_employee_number] += round(
                lamb_room_total_amount / lamb_room_window_total_hour * current_employee_work_hour_today, 2)

    # 海鲜
    row_number = 22
    sea_food_work_count_per_table = 1.05
    sea_food_window_total_hour = 0
    sea_food_total_amount = day_table_count * sea_food_work_count_per_table
    for number in range(1, 5):
        if daily_data["dish_prepare_work_hour"][f"sea_food{number}"] != "":
            current_employee_number = str(daily_data["dish_prepare_work_hour"][f"sea_food{number}"]).split(" ")[0]
            current_employee_name = str(daily_data["dish_prepare_work_hour"][f"sea_food{number}"]).split(" ", 1)[1]
            current_employee_work_hour_today = float(
                daily_data["dish_prepare_work_hour"][f"sea_food{number}_work_hour"])
            kitchen_work_count_xlsx_ws[f"A{row_number}"] = current_employee_number
            kitchen_work_count_xlsx_ws[f"B{row_number}"] = current_employee_name
            kitchen_work_count_xlsx_ws[f"C{row_number}"] = current_employee_work_hour_today
            sea_food_window_total_hour += current_employee_work_hour_today
            row_number += 1
    for number in range(1, 5):
        if daily_data["dish_prepare_work_hour"][f"sea_food{number}"] != "":
            current_employee_number = str(daily_data["dish_prepare_work_hour"][f"sea_food{number}"]).split(" ")[0]
            current_employee_work_hour_today = float(
                daily_data["dish_prepare_work_hour"][f"sea_food{number}_work_hour"])
            if not daily_kitchen_work_count.get(current_employee_number):
                daily_kitchen_work_count[current_employee_number] = 0
            daily_kitchen_work_count[current_employee_number] += round(
                sea_food_total_amount / sea_food_window_total_hour * current_employee_work_hour_today, 2)

    # 豆菌
    row_number = 30
    fungus_room_work_count_per_table = 1.2
    fungus_room_window_total_hour = 0
    fungus_room_total_amount = day_table_count * fungus_room_work_count_per_table
    for number in range(1, 5):
        if daily_data["dish_prepare_work_hour"][f"fungus_room{number}"] != "":
            current_employee_number = str(daily_data["dish_prepare_work_hour"][f"fungus_room{number}"]).split(" ")[0]
            current_employee_name = str(daily_data["dish_prepare_work_hour"][f"fungus_room{number}"]).split(" ", 1)[1]
            current_employee_work_hour_today = float(
                daily_data["dish_prepare_work_hour"][f"fungus_room{number}_work_hour"])
            kitchen_work_count_xlsx_ws[f"A{row_number}"] = current_employee_number
            kitchen_work_count_xlsx_ws[f"B{row_number}"] = current_employee_name
            kitchen_work_count_xlsx_ws[f"C{row_number}"] = current_employee_work_hour_today
            fungus_room_window_total_hour += current_employee_work_hour_today
            row_number += 1
    for number in range(1, 5):
        if daily_data["dish_prepare_work_hour"][f"fungus_room{number}"] != "":
            current_employee_number = str(daily_data["dish_prepare_work_hour"][f"fungus_room{number}"]).split(" ")[0]
            current_employee_work_hour_today = float(
                daily_data["dish_prepare_work_hour"][f"fungus_room{number}_work_hour"])
            if not daily_kitchen_work_count.get(current_employee_number):
                daily_kitchen_work_count[current_employee_number] = 0
            daily_kitchen_work_count[current_employee_number] += round(
                fungus_room_total_amount / fungus_room_window_total_hour * current_employee_work_hour_today, 2)

    # 粗加工
    row_number = 38
    rough_process_work_count_per_table = 0.95
    rough_process_window_total_hour = 0
    rough_process_total_amount = day_table_count * rough_process_work_count_per_table
    for number in range(1, 5):
        if daily_data["dish_prepare_work_hour"][f"rough_process{number}"] != "":
            current_employee_number = str(daily_data["dish_prepare_work_hour"][f"rough_process{number}"]).split(" ")[0]
            current_employee_name = str(daily_data["dish_prepare_work_hour"][f"rough_process{number}"]).split(" ", 1)[1]
            current_employee_work_hour_today = float(
                daily_data["dish_prepare_work_hour"][f"rough_process{number}_work_hour"])
            kitchen_work_count_xlsx_ws[f"A{row_number}"] = current_employee_number
            kitchen_work_count_xlsx_ws[f"B{row_number}"] = current_employee_name
            kitchen_work_count_xlsx_ws[f"C{row_number}"] = current_employee_work_hour_today
            rough_process_window_total_hour += current_employee_work_hour_today
            row_number += 1
    for number in range(1, 5):
        if daily_data["dish_prepare_work_hour"][f"rough_process{number}"] != "":
            current_employee_number = str(daily_data["dish_prepare_work_hour"][f"rough_process{number}"]).split(" ")[0]
            current_employee_work_hour_today = float(
                daily_data["dish_prepare_work_hour"][f"rough_process{number}_work_hour"])
            if not daily_kitchen_work_count.get(current_employee_number):
                daily_kitchen_work_count[current_employee_number] = 0
            daily_kitchen_work_count[current_employee_number] += round(
                rough_process_total_amount / rough_process_window_total_hour * current_employee_work_hour_today, 2)

    # 小吃房
    row_number = 46
    appetizer_work_count_per_table = 2.85
    appetizer_total_hour = 0
    appetizer_total_amount = day_table_count * appetizer_work_count_per_table
    for person in daily_data["appetizer_work_hour"]:
        current_employee_number = person.split(" ")[0]
        current_employee_name = person.split(" ", 1)[1]
        current_employee_work_hour_today = float(daily_data["appetizer_work_hour"][person]['worked_hour'])
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = current_employee_number
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = current_employee_name
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = current_employee_work_hour_today
        appetizer_total_hour += current_employee_work_hour_today
        row_number += 1
    for person in daily_data["appetizer_work_hour"]:
        current_employee_number = person.split(" ")[0]
        current_employee_work_hour_today = float(daily_data["appetizer_work_hour"][person]['worked_hour'])
        if not daily_kitchen_work_count.get(current_employee_number):
            daily_kitchen_work_count[current_employee_number] = 0
        daily_kitchen_work_count[current_employee_number] += round(
            appetizer_total_amount / appetizer_total_hour * current_employee_work_hour_today, 2)

    # 结束计算，关闭文件
    kitchen_work_count_xlsx_wb.save(kitchen_work_count_xlsx_file_path)
    kitchen_work_count_xlsx_wb.close()
    daily_data["close_data"]["daily_kitchen_work_count"] = daily_kitchen_work_count
    data_handler.save_daily_data_as_json(json.dumps(daily_data))


def daily_tips_count():
    employee_data = data_handler.get_employee_data()
    daily_data = data_handler.get_daily_data()
    daily_tips_count_data = {}
    date_next_day = data_handler.get_date_next_day(data_handler.get_server_date())
    daily_data_next_day = data_handler.get_daily_data(date_next_day)
    tips_count_xlsx_file_path = ""
    for file in os.listdir(destination_path):
        if "小费计算表格" in file:
            tips_count_xlsx_file_path = destination_path + "/" + file
            break
    time.sleep(1)
    tips_count_xlsx_wb = openpyxl.load_workbook(tips_count_xlsx_file_path)
    tips_count_xlsx_ws = tips_count_xlsx_wb.active

    # 日期
    tips_count_xlsx_ws['A1'] = data_handler.get_server_date()

    # 服务组小费
    server_group_row_length = 15
    daily_total_tips = 0
    starting_row = 4
    server_department_tips_ratio = 0.47
    for team in daily_data["server_group_info"]:
        team_total_tips = 0
        try:
            team_total_tips += float(daily_data["server_group_info"][team]["cash_tips"])
        except:
            pass
        try:
            team_total_tips += float(daily_data["server_group_info"][team]["card_tips"])
        except:
            pass
        try:
            team_total_tips += float(daily_data["server_group_info"][team]["alipay_tips"])
        except:
            pass
        try:
            team_total_tips += float(daily_data["server_group_info"][team]["wechatpay_tips"])
        except:
            pass
        transfer_tips_total = 0
        try:
            transfer_tips_total += float(daily_data["server_group_info"][team]["transfer_cash_tips"])
        except:
            pass
        try:
            transfer_tips_total += float(daily_data["server_group_info"][team]["transfer_card_tips"])
        except:
            pass
        try:
            transfer_tips_total += float(daily_data["server_group_info"][team]["transfer_alipay_tips"])
        except:
            pass
        try:
            transfer_tips_total += float(daily_data["server_group_info"][team]["transfer_wechatpay_tips"])
        except:
            pass
        daily_total_tips += team_total_tips + transfer_tips_total
        if daily_data["server_group_info"][team].get("has_transfer_group") and daily_data["server_group_info"][team][
            "has_transfer_group"] == "True" and transfer_tips_total != 0:
            server_team_members_info = {}
            team_total_ratios = 0
            current_team_total_tips = team_total_tips + transfer_tips_total / 2
            tips_count_xlsx_ws[f'C{starting_row}'] = current_team_total_tips
            current_row = starting_row
            for item in daily_data["server_group_info"][team]:
                if "team_member" in item and "work_hour" not in item and daily_data["server_group_info"][team][
                    item] != "":
                    current_team_member = daily_data["server_group_info"][team][item]
                    employee_id = str(current_team_member).split(" ", 1)[0]
                    employee_name = str(current_team_member).split(" ", 1)[1]
                    employee_tips_ratio = float(
                        employee_data[str(current_team_member).split(" ")[0]]["server_tips_ratio"])
                    employee_work_hour_ratio = float(daily_data["server_group_info"][team][item + "_work_hour"]) / 8
                    employee_tips_ratio_today = employee_tips_ratio * employee_work_hour_ratio
                    tips_count_xlsx_ws[f'H{current_row}'] = employee_work_hour_ratio
                    tips_count_xlsx_ws[f'A{current_row}'] = employee_id
                    tips_count_xlsx_ws[f'B{current_row}'] = employee_name
                    try:
                        tips_count_xlsx_ws[f'G{current_row}'] = float(employee_tips_ratio_today)
                    except:
                        tips_count_xlsx_ws[f'G{current_row}'] = 0
                    server_team_members_info[employee_id] = {
                        "employee_tips_ratio_today": employee_tips_ratio_today
                    }
                    team_total_ratios += employee_tips_ratio_today
                    current_row += 1

            for employee in server_team_members_info:
                current_employee_tips_ratio_today = server_team_members_info[employee]["employee_tips_ratio_today"]
                if not daily_tips_count_data.get(employee):
                    daily_tips_count_data[employee] = 0
                daily_tips_count_data[employee] += round(
                    current_team_total_tips * server_department_tips_ratio / team_total_ratios * current_employee_tips_ratio_today)
            starting_row += server_group_row_length

            # 交接小组
            current_row = starting_row
            server_team_members_info = {}
            team_total_ratios = 0
            current_team_total_tips = transfer_tips_total / 2
            tips_count_xlsx_ws[f'C{starting_row}'] = current_team_total_tips
            for item in daily_data["server_group_info"][team]:
                if "transfer_member" in item and "work_hour" not in item and daily_data["server_group_info"][team][
                    item] != "":
                    current_team_member = daily_data["server_group_info"][team][item]
                    employee_id = str(current_team_member).split(" ", 1)[0]
                    employee_name = str(current_team_member).split(" ", 1)[1]
                    employee_tips_ratio = float(
                        employee_data[str(current_team_member).split(" ")[0]]["server_tips_ratio"])
                    employee_work_hour_ratio = float(daily_data["server_group_info"][team][item + "_work_hour"]) / 8
                    employee_tips_ratio_today = employee_tips_ratio * employee_work_hour_ratio
                    tips_count_xlsx_ws[f'H{current_row}'] = employee_work_hour_ratio
                    tips_count_xlsx_ws[f'A{current_row}'] = employee_id
                    tips_count_xlsx_ws[f'B{current_row}'] = employee_name
                    try:
                        tips_count_xlsx_ws[f'G{current_row}'] = float(employee_tips_ratio_today)
                    except:
                        tips_count_xlsx_ws[f'G{current_row}'] = 0
                    server_team_members_info[employee_id] = {
                        "employee_tips_ratio_today": employee_tips_ratio_today
                    }
                    team_total_ratios += employee_tips_ratio_today
                    current_row += 1

            for employee in server_team_members_info:
                current_employee_tips_ratio_today = server_team_members_info[employee]["employee_tips_ratio_today"]
                if not daily_tips_count_data.get(employee):
                    daily_tips_count_data[employee] = 0
                daily_tips_count_data[employee] += round(
                    current_team_total_tips * server_department_tips_ratio / team_total_ratios * current_employee_tips_ratio_today)
            starting_row += server_group_row_length
        else:
            server_team_members_info = {}
            team_total_ratios = 0
            current_team_total_tips = team_total_tips
            tips_count_xlsx_ws[f'C{starting_row}'] = current_team_total_tips
            current_row = starting_row
            for item in daily_data["server_group_info"][team]:
                if "team_member" in item and "work_hour" not in item and \
                        daily_data["server_group_info"][team][item] != "":
                    current_team_member = daily_data["server_group_info"][team][item]
                    employee_id = str(current_team_member).split(" ", 1)[0]
                    employee_name = str(current_team_member).split(" ", 1)[1]
                    employee_tips_ratio = float(
                        employee_data[str(current_team_member).split(" ")[0]]["server_tips_ratio"])
                    employee_work_hour_ratio = float(
                        daily_data["server_group_info"][team][item + "_work_hour"]) / 8
                    employee_tips_ratio_today = employee_tips_ratio * employee_work_hour_ratio
                    tips_count_xlsx_ws[f'H{current_row}'] = employee_work_hour_ratio
                    tips_count_xlsx_ws[f'A{current_row}'] = employee_id
                    tips_count_xlsx_ws[f'B{current_row}'] = employee_name
                    try:
                        tips_count_xlsx_ws[f'G{current_row}'] = float(employee_tips_ratio_today)
                    except:
                        tips_count_xlsx_ws[f'G{current_row}'] = 0
                    server_team_members_info[employee_id] = {
                        "employee_tips_ratio_today": employee_tips_ratio_today
                    }
                    team_total_ratios += employee_tips_ratio_today
                    current_row += 1

            for employee in server_team_members_info:
                current_employee_tips_ratio_today = server_team_members_info[employee]["employee_tips_ratio_today"]
                if not daily_tips_count_data.get(employee):
                    daily_tips_count_data[employee] = 0
                daily_tips_count_data[
                    employee] += round(
                    current_team_total_tips * server_department_tips_ratio / team_total_ratios * current_employee_tips_ratio_today)
            starting_row += server_group_row_length

    # 门迎小费
    starting_row = 158
    host_department_ratio = 0.12
    host_department_total_ratio = 0
    host_employee_info_today = {}
    host_deducted_tips = 0
    for item in daily_data["general_data"]["tips_give_out_from_department"]:
        if daily_data["general_data"]["tips_give_out_from_department"][item]["deduct_from_department"] == "门迎":
            host_deducted_tips -= float(daily_data["general_data"]["tips_give_out_from_department"][item]["amount"])
    for person in daily_data["host_work_hour"]:
        employee_id = str(person).split(" ", 1)[0]
        employee_name = str(person).split(" ", 1)[1]
        employee_tips_ratio = float(employee_data[str(person).split(" ")[0]]["host_tips_ratios"])
        employee_work_hour = float(daily_data["host_work_hour"][person]['worked_hour'])
        employee_ratio_today = employee_tips_ratio * employee_work_hour
        tips_count_xlsx_ws[f'A{starting_row}'] = employee_id
        tips_count_xlsx_ws[f'B{starting_row}'] = employee_name
        try:
            tips_count_xlsx_ws[f'C{starting_row}'] = employee_tips_ratio
        except:
            tips_count_xlsx_ws[f'C{starting_row}'] = 0
        tips_count_xlsx_ws[f'E{starting_row}'] = employee_work_hour
        host_department_total_ratio += employee_ratio_today
        host_employee_info_today[employee_id] = {
            "employee_ratio_today": employee_ratio_today
        }
        starting_row += 1

    for employee in host_employee_info_today:
        current_employee_tips_ratio_today = host_employee_info_today[employee]["employee_ratio_today"]
        if not daily_tips_count_data.get(employee):
            daily_tips_count_data[employee] = 0
        daily_tips_count_data[
            employee] += round((
                                       daily_total_tips * host_department_ratio - host_deducted_tips) / host_department_total_ratio * current_employee_tips_ratio_today)

    # 保洁小费
    starting_row = 179
    cleaner_department_ratio = 0.02
    cleaner_department_total_ratio = 0
    cleaner_employee_info_today = {}
    cleaner_deducted_tips = 0
    for item in daily_data["general_data"]["tips_give_out_from_department"]:
        if daily_data["general_data"]["tips_give_out_from_department"][item]["deduct_from_department"] == "保洁":
            cleaner_deducted_tips -= float(daily_data["general_data"]["tips_give_out_from_department"][item]["amount"])
    for person in daily_data["cleaner_work_hour"]:
        employee_id = str(person).split(" ", 1)[0]
        employee_name = str(person).split(" ", 1)[1]
        employee_tip_ratio = float(employee_data[str(person).split(" ")[0]]["cleaner_tips_ratio"])
        employee_work_hour = float(daily_data["cleaner_work_hour"][person])
        employee_ratio_today = employee_tip_ratio * employee_work_hour
        tips_count_xlsx_ws[f'A{starting_row}'] = employee_id
        tips_count_xlsx_ws[f'B{starting_row}'] = employee_name
        try:
            tips_count_xlsx_ws[f'C{starting_row}'] = employee_tip_ratio
        except:
            tips_count_xlsx_ws[f'C{starting_row}'] = 0
        tips_count_xlsx_ws[f'E{starting_row}'] = employee_work_hour
        cleaner_department_total_ratio += employee_ratio_today
        cleaner_employee_info_today[employee_id] = {
            "employee_ratio_today": employee_ratio_today
        }
        starting_row += 1

    for employee in cleaner_employee_info_today:
        current_employee_tips_ratio_today = cleaner_employee_info_today[employee]["employee_ratio_today"]
        if not daily_tips_count_data.get(employee):
            daily_tips_count_data[employee] = 0
        daily_tips_count_data[
            employee] += round((
                                       daily_total_tips * cleaner_department_ratio - cleaner_deducted_tips) / cleaner_department_total_ratio * current_employee_tips_ratio_today)

    # 管理及特岗小费
    # 大堂经理
    starting_row = 192
    front_manager = daily_data["manager_daily_work_info_info"]["front_manager"]
    front_manager_id = str(front_manager).split(" ", 1)[0]
    front_manager_name = str(front_manager).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A{starting_row}'] = front_manager_id
    tips_count_xlsx_ws[f'B{starting_row}'] = front_manager_name
    front_manager_department_tips_ratio = 0.0283
    front_manager_today_tips = daily_total_tips * front_manager_department_tips_ratio
    if not daily_tips_count_data.get(front_manager_id):
        daily_tips_count_data[front_manager_id] = 0
    daily_tips_count_data[front_manager_id] += round(front_manager_today_tips)

    # 值班经理
    starting_row = 193
    duty_manager = daily_data["manager_daily_work_info_info"]["duty_manager"]
    duty_manager_id = str(duty_manager).split(" ", 1)[0]
    duty_manager_name = str(duty_manager).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A{starting_row}'] = duty_manager_id
    tips_count_xlsx_ws[f'B{starting_row}'] = duty_manager_name
    duty_manager_department_tips_ratio = 0.0283
    duty_manager_today_tips = daily_total_tips * duty_manager_department_tips_ratio
    if not daily_tips_count_data.get(duty_manager_id):
        daily_tips_count_data[duty_manager_id] = 0
    daily_tips_count_data[duty_manager_id] += round(duty_manager_today_tips)

    # 后堂经理
    starting_row = 194
    kitchen_manager = daily_data["manager_daily_work_info_info"]["kitchen_manager"]
    kitchen_manager_id = str(kitchen_manager).split(" ", 1)[0]
    kitchen_manager_name = str(kitchen_manager).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A{starting_row}'] = kitchen_manager_id
    tips_count_xlsx_ws[f'B{starting_row}'] = kitchen_manager_name
    kitchen_manager_department_tips_ratio = 0.0283
    kitchen_manager_today_tips = daily_total_tips * kitchen_manager_department_tips_ratio
    if not daily_tips_count_data.get(kitchen_manager_id):
        daily_tips_count_data[kitchen_manager_id] = 0
    daily_tips_count_data[kitchen_manager_id] += round(kitchen_manager_today_tips)

    # 变脸师
    starting_row = 196
    face_changing_master = daily_data["manager_daily_work_info_info"]["face_changing_master"]
    face_changing_master_id = str(face_changing_master).split(" ", 1)[0]
    face_changing_master_name = str(face_changing_master).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A{starting_row}'] = face_changing_master_id
    tips_count_xlsx_ws[f'B{starting_row}'] = face_changing_master_name
    face_changing_master_department_tips_ratio = 0.005
    face_changing_master_today_tips = daily_total_tips * face_changing_master_department_tips_ratio
    if not daily_tips_count_data.get(face_changing_master_id):
        daily_tips_count_data[face_changing_master_id] = 0
    daily_tips_count_data[face_changing_master_id] += round(face_changing_master_today_tips)

    # 员工餐师傅
    starting_row = 197
    employee_meal_cooker = daily_data["manager_daily_work_info_info"]["employee_meal_cooker"]
    employee_meal_cooker_id = str(employee_meal_cooker).split(" ", 1)[0]
    employee_meal_cooker_name = str(employee_meal_cooker).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A{starting_row}'] = employee_meal_cooker_id
    tips_count_xlsx_ws[f'B{starting_row}'] = employee_meal_cooker_name
    employee_meal_cooker_department_tips_ratio = 0.015
    employee_meal_cooker_today_tips = daily_total_tips * employee_meal_cooker_department_tips_ratio
    if not daily_tips_count_data.get(employee_meal_cooker_id):
        daily_tips_count_data[employee_meal_cooker_id] = 0
    daily_tips_count_data[employee_meal_cooker_id] += round(employee_meal_cooker_today_tips)

    # 文员
    starting_row = 198
    clerk = daily_data["manager_daily_work_info_info"]["clerk"]
    clerk_id = str(clerk).split(" ", 1)[0]
    clerk_name = str(clerk).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A{starting_row}'] = clerk_id
    tips_count_xlsx_ws[f'B{starting_row}'] = clerk_name
    clerk_department_tips_ratio = 0.005
    clerk_today_tips = daily_total_tips * clerk_department_tips_ratio
    if not daily_tips_count_data.get(clerk_id):
        daily_tips_count_data[clerk_id] = 0
    daily_tips_count_data[clerk_id] += round(clerk_today_tips)

    # 质检员
    starting_row = 199
    quality_inspector = daily_data["manager_daily_work_info_info"]["quality_inspector"]
    quality_inspector_id = str(quality_inspector).split(" ", 1)[0]
    quality_inspector_name = str(quality_inspector).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A{starting_row}'] = quality_inspector_id
    tips_count_xlsx_ws[f'B{starting_row}'] = quality_inspector_name
    quality_inspector_department_tips_ratio = 0.005
    quality_inspector_today_tips = daily_total_tips * quality_inspector_department_tips_ratio
    if not daily_tips_count_data.get(quality_inspector_id):
        daily_tips_count_data[quality_inspector_id] = 0
    daily_tips_count_data[quality_inspector_id] += round(quality_inspector_today_tips)

    # 报销员
    starting_row = 200
    expense_accountant = daily_data["manager_daily_work_info_info"]["expense_accountant"]
    expense_accountant_id = str(expense_accountant).split(" ", 1)[0]
    expense_accountant_name = str(expense_accountant).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A{starting_row}'] = expense_accountant_id
    tips_count_xlsx_ws[f'B{starting_row}'] = expense_accountant_name
    expense_accountant_department_tips_ratio = 0.005
    expense_accountant_today_tips = daily_total_tips * expense_accountant_department_tips_ratio
    if not daily_tips_count_data.get(expense_accountant_id):
        daily_tips_count_data[expense_accountant_id] = 0
    daily_tips_count_data[expense_accountant_id] += round(expense_accountant_today_tips)

    # 上菜口小费
    starting_row = 206
    dish_prepare_department_ratio = 0.092
    dish_prepare_department_total_ratio = 0
    dish_prepare_employee_info_today = {}
    dish_prepare_deducted_tips = 0
    for item in daily_data["general_data"]["tips_give_out_from_department"]:
        if daily_data["general_data"]["tips_give_out_from_department"][item]["deduct_from_department"] == "上菜":
            dish_prepare_deducted_tips -= float(
                daily_data["general_data"]["tips_give_out_from_department"][item]["amount"])
    for item in daily_data["dish_prepare_work_hour"]:
        if "work_hour" not in item and daily_data["dish_prepare_work_hour"][item] != "":
            employee_id = str(daily_data["dish_prepare_work_hour"][item]).split(" ", 1)[0]
            employee_name = str(daily_data["dish_prepare_work_hour"][item]).split(" ", 1)[1]
            employee_tip_ratio = float(employee_data[str(daily_data["dish_prepare_work_hour"][item]).split(" ", 1)[0]][
                                           "dish_prepare_tips_ratio"])
            employee_work_hour = float(daily_data["dish_prepare_work_hour"][item + "_work_hour"])
            employee_ratio_today = employee_tip_ratio * employee_work_hour
            tips_count_xlsx_ws[f'A{starting_row}'] = employee_id
            tips_count_xlsx_ws[f'B{starting_row}'] = employee_name
            try:
                tips_count_xlsx_ws[f'C{starting_row}'] = employee_tip_ratio
            except:
                tips_count_xlsx_ws[f'C{starting_row}'] = 0
            tips_count_xlsx_ws[f'E{starting_row}'] = employee_work_hour
            dish_prepare_department_total_ratio += employee_ratio_today
            dish_prepare_employee_info_today[employee_id] = {
                "employee_ratio_today": employee_ratio_today
            }
            starting_row += 1

    for employee in dish_prepare_employee_info_today:
        current_employee_tips_ratio_today = dish_prepare_employee_info_today[employee]["employee_ratio_today"]
        if not daily_tips_count_data.get(employee):
            daily_tips_count_data[employee] = 0
        daily_tips_count_data[
            employee] += round((
                                       daily_total_tips * dish_prepare_department_ratio - dish_prepare_deducted_tips) / dish_prepare_department_total_ratio * current_employee_tips_ratio_today)

    # 小吃房小费
    starting_row = 227
    appetizer_department_ratio = 0.047
    appetizer_department_total_ratio = 0
    appetizer_employee_info_today = {}
    appetizer_deducted_tips = 0
    for item in daily_data["general_data"]["tips_give_out_from_department"]:
        if daily_data["general_data"]["tips_give_out_from_department"][item]["deduct_from_department"] == "小吃":
            appetizer_deducted_tips -= float(
                daily_data["general_data"]["tips_give_out_from_department"][item]["amount"])
    for person in daily_data["appetizer_work_hour"]:
        employee_id = person.split(" ", 1)[0]
        employee_name = person.split(" ", 1)[1]
        print(person)
        employee_tip_ratio = float(employee_data[str(person.split(" ", 1)[0])]["appetizer_tips_ratio"])
        employee_work_hour = float(daily_data["appetizer_work_hour"][person]['worked_hour'])
        employee_ratio_today = employee_tip_ratio * employee_work_hour
        tips_count_xlsx_ws[f'A{starting_row}'] = employee_id
        tips_count_xlsx_ws[f'B{starting_row}'] = employee_name
        try:
            tips_count_xlsx_ws[f'C{starting_row}'] = float(employee_tip_ratio)
        except:
            tips_count_xlsx_ws[f'C{starting_row}'] = 0
        tips_count_xlsx_ws[f'E{starting_row}'] = float(employee_work_hour)
        appetizer_department_total_ratio += employee_ratio_today
        appetizer_employee_info_today[employee_id] = {
            "employee_ratio_today": employee_ratio_today
        }
        starting_row += 1

    for employee in appetizer_employee_info_today:
        current_employee_tips_ratio_today = appetizer_employee_info_today[employee]["employee_ratio_today"]
        if not daily_tips_count_data.get(employee):
            daily_tips_count_data[employee] = 0
        daily_tips_count_data[
            employee] += round((
                                       daily_total_tips * appetizer_department_ratio - appetizer_deducted_tips) / appetizer_department_total_ratio * current_employee_tips_ratio_today)

    # 捞面小费
    starting_row = 252
    noodle_dance_department_ratio = 0.017
    noodle_dance_department_total_ratio = 0
    noodle_dance_employee_info_today = {}
    noodle_dance_deducted_tips = 0
    for item in daily_data["general_data"]["tips_give_out_from_department"]:
        if daily_data["general_data"]["tips_give_out_from_department"][item]["deduct_from_department"] == "捞面":
            noodle_dance_deducted_tips -= float(
                daily_data["general_data"]["tips_give_out_from_department"][item]["amount"])
    for person in daily_data["noodle_dance_work_hour"]:
        employee_id = person.split(" ", 1)[0]
        employee_name = person.split(" ", 1)[1]
        employee_tip_ratio = float(employee_data[str(person.split(" ", 1)[0])]["noodle_dance_tips_ratios"])
        employee_work_hour = float(daily_data["noodle_dance_work_hour"][person])
        employee_ratio_today = employee_tip_ratio * employee_work_hour
        tips_count_xlsx_ws[f'A{starting_row}'] = employee_id
        tips_count_xlsx_ws[f'B{starting_row}'] = employee_name
        try:
            tips_count_xlsx_ws[f'C{starting_row}'] = float(employee_tip_ratio)
        except:
            tips_count_xlsx_ws[f'C{starting_row}'] = 0
        tips_count_xlsx_ws[f'E{starting_row}'] = float(employee_work_hour)
        noodle_dance_department_total_ratio += employee_ratio_today
        noodle_dance_employee_info_today[employee_id] = {
            "employee_ratio_today": employee_ratio_today
        }
        starting_row += 1

    for employee in noodle_dance_employee_info_today:
        current_employee_tips_ratio_today = noodle_dance_employee_info_today[employee]["employee_ratio_today"]
        if not daily_tips_count_data.get(employee):
            daily_tips_count_data[employee] = 0
        daily_tips_count_data[
            employee] += round((
                                       daily_total_tips * noodle_dance_department_ratio - noodle_dance_deducted_tips) / noodle_dance_department_total_ratio * current_employee_tips_ratio_today)

    # 传配小费
    starting_row = 264
    dish_runner_department_ratio = 0.082
    dish_runner_department_total_ratio = 0
    dish_runner_employee_info_today = {}
    dish_runner_deducted_tips = 0
    for item in daily_data["general_data"]["tips_give_out_from_department"]:
        if daily_data["general_data"]["tips_give_out_from_department"][item]["deduct_from_department"] == "传配":
            dish_runner_deducted_tips -= float(
                daily_data["general_data"]["tips_give_out_from_department"][item]["amount"])
    for person in daily_data["dish_runner_work_hour"]:
        employee_id = person.split(" ", 1)[0]
        employee_name = person.split(" ", 1)[1]
        employee_tip_ratio = float(employee_data[str(person.split(" ", 1)[0])]["dish_runner_tips_ratio"])
        employee_work_hour = float(daily_data["dish_runner_work_hour"][person]["morning_worked_hour"]) + float(
            daily_data["dish_runner_work_hour"][person]["evening_worked_hour"])
        employee_ratio_today = employee_tip_ratio * employee_work_hour
        tips_count_xlsx_ws[f'A{starting_row}'] = employee_id
        tips_count_xlsx_ws[f'B{starting_row}'] = employee_name
        try:
            tips_count_xlsx_ws[f'C{starting_row}'] = float(employee_tip_ratio)
        except:
            tips_count_xlsx_ws[f'C{starting_row}'] = 0
        tips_count_xlsx_ws[f'E{starting_row}'] = float(employee_work_hour)
        dish_runner_department_total_ratio += employee_ratio_today
        dish_runner_employee_info_today[employee_id] = {
            "employee_ratio_today": employee_ratio_today
        }
        starting_row += 1

    for employee in dish_runner_employee_info_today:
        current_employee_tips_ratio_today = dish_runner_employee_info_today[employee]["employee_ratio_today"]
        if not daily_tips_count_data.get(employee):
            daily_tips_count_data[employee] = 0
        daily_tips_count_data[
            employee] += round((
                                       daily_total_tips * dish_runner_department_ratio - dish_runner_deducted_tips) / dish_runner_department_total_ratio * current_employee_tips_ratio_today)

    # 部门小费给出 以及 小费扣补
    starting_row = 313
    for item in daily_data["general_data"]["tips_give_out_from_department"]:
        if not daily_tips_count_data.get(
                str(daily_data["general_data"]["tips_give_out_from_department"][item]['employee_name']).split(" ")[0]):
            daily_tips_count_data[
                str(daily_data["general_data"]["tips_give_out_from_department"][item]['employee_name']).split(" ")[
                    0]] = 0
        daily_tips_count_data[
            str(daily_data["general_data"]["tips_give_out_from_department"][item]['employee_name']).split(" ")[
                0]] += round(float(daily_data["general_data"]["tips_give_out_from_department"][item]['amount']))
        tips_count_xlsx_ws[f'A{starting_row}'] = \
            str(daily_data["general_data"]["tips_give_out_from_department"][item]['employee_name']).split(" ")[0]
        tips_count_xlsx_ws[f'B{starting_row}'] = \
            str(daily_data["general_data"]["tips_give_out_from_department"][item]['employee_name']).split(" ", 1)[1]
        tips_count_xlsx_ws[f'C{starting_row}'] = \
            daily_data["general_data"]["tips_give_out_from_department"][item]['reason']
        tips_count_xlsx_ws[f'H{starting_row}'] = \
            daily_data["general_data"]["tips_give_out_from_department"][item]['deduct_from_department']
        tips_count_xlsx_ws[f'I{starting_row}'] = \
            float(daily_data["general_data"]["tips_give_out_from_department"][item]['amount'])
        starting_row += 1
        if daily_data["general_data"]["tips_give_out_from_department"][item]["long_term"] == "yes":
            if not daily_data_next_day.get("general_data"):
                daily_data_next_day["general_data"] = {}
            if not daily_data_next_day["general_data"].get("tips_give_out_from_department"):
                daily_data_next_day["general_data"]["tips_give_out_from_department"] = {}
            daily_data_next_day["general_data"]["tips_give_out_from_department"][item] = \
                daily_data["general_data"]["tips_give_out_from_department"][item]

    for item in daily_data["general_data"]["tips_adjustment"]:
        if not daily_tips_count_data.get(
                str(daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ")[0]):
            daily_tips_count_data[
                str(daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ")[0]] = 0

        if float(daily_data["general_data"]["tips_adjustment"][item]["amount"]) >= 0:
            daily_tips_count_data[
                str(daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ")[0]] += float(
                daily_data["general_data"]["tips_adjustment"][item]["amount"])
            tips_count_xlsx_ws[f'A{starting_row}'] = \
                str(daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ")[0]
            tips_count_xlsx_ws[f'B{starting_row}'] = \
                str(daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ", 1)[1]
            tips_count_xlsx_ws[f'C{starting_row}'] = \
                daily_data["general_data"]["tips_adjustment"][item]['reason']
            tips_count_xlsx_ws[f'H{starting_row}'] = "无"
            tips_count_xlsx_ws[f'I{starting_row}'] = \
                float(daily_data["general_data"]["tips_adjustment"][item]['amount'])
            starting_row += 1
        else:
            if not daily_tips_count_data[
                       str(daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ")[0]] + float(
                daily_data["general_data"]["tips_adjustment"][item]["amount"]) < 0:
                daily_tips_count_data[
                    str(daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ")[0]] += float(
                    daily_data["general_data"]["tips_adjustment"][item]["amount"])

                tips_count_xlsx_ws[f'A{starting_row}'] = \
                    str(daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ")[0]
                tips_count_xlsx_ws[f'B{starting_row}'] = \
                    str(daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ", 1)[1]
                tips_count_xlsx_ws[f'C{starting_row}'] = \
                    daily_data["general_data"]["tips_adjustment"][item]['reason']
                tips_count_xlsx_ws[f'H{starting_row}'] = "无"
                tips_count_xlsx_ws[f'I{starting_row}'] = \
                    float(daily_data["general_data"]["tips_adjustment"][item]['amount'])
                starting_row += 1
            else:
                if not daily_data_next_day.get("general_data"):
                    daily_data_next_day["general_data"] = {}
                if not daily_data_next_day["general_data"].get("tips_adjustment"):
                    daily_data_next_day["general_data"]["tips_adjustment"] = {}

                tips_count_xlsx_ws[f'A{starting_row}'] = \
                    str(daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ")[0]
                tips_count_xlsx_ws[f'B{starting_row}'] = \
                    str(daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ", 1)[1]
                tips_count_xlsx_ws[f'C{starting_row}'] = \
                    daily_data["general_data"]["tips_adjustment"][item]['reason']
                tips_count_xlsx_ws[f'H{starting_row}'] = "无"
                tips_count_xlsx_ws[f'I{starting_row}'] = -float(daily_tips_count_data[str(
                    daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ")[0]])

                daily_data_next_day["general_data"]["tips_adjustment"][item] = \
                    daily_data["general_data"]["tips_adjustment"][item]
                daily_data_next_day["general_data"]["tips_adjustment"][item]['amount'] = float(
                    daily_data["general_data"]["tips_adjustment"][item]["amount"]) + daily_tips_count_data[str(
                    daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ")[0]]
                daily_tips_count_data[
                    str(daily_data["general_data"]["tips_adjustment"][item]['employee_name']).split(" ")[0]] = 0
                starting_row += 1

    tips_count_xlsx_wb.save(tips_count_xlsx_file_path)
    tips_count_xlsx_wb.close()
    daily_data["close_data"]["daily_tips_count"] = daily_tips_count_data
    data_handler.save_daily_data_as_json(json.dumps(daily_data))
    data_handler.save_daily_data_as_json(json.dumps(daily_data_next_day), date=date_next_day)


def daily_take_out_bonus_count():
    employee_data = data_handler.get_employee_data()
    daily_data = data_handler.get_daily_data()
    daily_takeout_data = daily_data["take_out_data"]
    daily_take_out_bonus_data = {}
    for group in daily_takeout_data:
        current_group = daily_takeout_data[group]
        group_total_hour = 0
        group_total_money = 0
        current_group_members = {}
        try:
            group_total_money -= float(current_group["fantuan_fee"]) * 0.05
        except:
            pass
        try:
            group_total_money += float(current_group["fantuan_pos_amount"]) * 0.05
        except:
            pass
        try:
            group_total_money += float(current_group["uber_paid_amount"]) * 0.05
        except:
            pass
        try:
            group_total_money += float(current_group["Snappy_amount"]) * 0.05
        except:
            pass
        try:
            group_total_money += float(current_group["cash_takeout"]) * 0.05
        except:
            pass
        try:
            group_total_money += float(current_group["card_take_out"]) * 0.05
        except:
            pass
        try:
            group_total_money += float(current_group["tips"])
        except:
            pass
        for item in current_group:
            if "team_member" in item and "work_hour" not in item and current_group[item] != "":
                team_member = str(current_group[item])
                team_member_id = team_member.split(" ")[0]
                team_member_work_hour = float(current_group[f'{item}_work_hour'])
                group_total_hour += team_member_work_hour
                current_group_members[team_member_id] = {"team_member_work_hour": team_member_work_hour}

        for member in current_group_members:
            if not daily_take_out_bonus_data.get(member):
                daily_take_out_bonus_data[member] = 0
            daily_take_out_bonus_data[member] += group_total_money / group_total_hour * float(
                current_group_members[member]["team_member_work_hour"])
    daily_data["close_data"]["daily_take_out_bonus_data"] = daily_take_out_bonus_data
    data_handler.save_daily_data_as_json(json.dumps(daily_data))


def get_haidilao_pos_data():
    s = requests.Session()
    # login, and get my personal info
    haidilao_pos_user_name = "10312229"
    haidilao_pos_password = "6/9USW7DQhgs+h1inlAmXRX+q9oVt2BkL9qp1LtG5og="

    proxies = {
        "https": "2.1.22.183",
        "http": "2.1.22.183"
    }

    print("login, and get my personal info")
    payload = {
        "userName": haidilao_pos_user_name,
        "password": haidilao_pos_password,
        "osId": 0,
        "type": "web",
        # "sig": "7b909d5d8e05f146bb89f81fc661c481"
    }
    pos_res_info = s.post("https://pos.haidilao.com:8036/login", json=payload, proxies=proxies)  # this is raw info
    json_pos_res_info = json.loads(pos_res_info.content)  # this is result converted raw info to json info
    s.headers.update(
        {"token": json.loads(pos_res_info.content)["obj"]["Token"]})  # add temperary token into session header

    # import daily summary data to daily close book
    params = {
        "groupCollect1": 1,
        "date": f"{data_handler.get_server_date()[0:4]}-{data_handler.get_server_date()[4:6]}-{data_handler.get_server_date()[6:8]},{data_handler.get_server_date()[0:4]}-{data_handler.get_server_date()[4:6]}-{data_handler.get_server_date()[6:8]}",
        "pageSize": 30,
        "pageNum": 1,
        "shopId": json_pos_res_info["obj"]["shopList"][0]["id"],
        "userName": json_pos_res_info["obj"]["userName"],
        "userId": json_pos_res_info["obj"]["userId"],
    }
    daily_summary_data_url = "https://pos.haidilao.com:8032/repBusinessTable/list"
    daily_summary_data_result = s.get(daily_summary_data_url, params=params, proxies=proxies)
    json_daily_summary_data_result = json.loads(daily_summary_data_result.content)

    morning_table_count = json_daily_summary_data_result["data"]["list"][0]["wuTableNum"]
    evening_table_count = json_daily_summary_data_result["data"]["list"][0]["wanTableNum"]

    dine_in_money = json_daily_summary_data_result["data"]["list"][0]["payShop"]
    take_out_money = json_daily_summary_data_result["data"]["list"][0]["paySend"]
    retail_money = json_daily_summary_data_result["data"]["list"][0]["payOut"]

    # put data into file
    # close book file
    close_book_xlsx_file_path = ""
    for file in os.listdir(data_handler.get_current_daily_file_path() + f"/{data_handler.get_server_date()}收账明细"):
        if "收账本" in file:
            close_book_xlsx_file_path = data_handler.get_current_daily_file_path() + f"/{data_handler.get_server_date()}收账明细/" + file
            break
    time.sleep(1)
    close_book_count_xlsx_wb = openpyxl.load_workbook(close_book_xlsx_file_path)
    close_book_count_xlsx_ws = close_book_count_xlsx_wb.active

    close_book_count_xlsx_ws['B1'] = float(retail_money)
    close_book_count_xlsx_ws['B2'] = float(dine_in_money)
    close_book_count_xlsx_ws['B3'] = float(take_out_money)

    close_book_count_xlsx_wb.save(close_book_xlsx_file_path)
    close_book_count_xlsx_wb.close()

    # dish runner file
    busser_work_count_xlsx_file_path = ""
    for file in os.listdir(data_handler.get_current_daily_file_path()):
        if "传配组打卡计件" in file:
            busser_work_count_xlsx_file_path = data_handler.get_current_daily_file_path() + "/" + file
            break
    time.sleep(1)
    busser_work_count_xlsx_wb = openpyxl.load_workbook(busser_work_count_xlsx_file_path)
    busser_work_count_xlsx_ws = busser_work_count_xlsx_wb.active

    busser_work_count_xlsx_ws["B3"] = int(morning_table_count)
    busser_work_count_xlsx_ws["B4"] = int(evening_table_count)
    busser_work_count_xlsx_wb.save(busser_work_count_xlsx_file_path)
    busser_work_count_xlsx_wb.close()

    # kitchen file
    kitchen_work_count_xlsx_file_path = ""
    for file in os.listdir(data_handler.get_current_daily_file_path()):
        if "厨房计件" in file:
            kitchen_work_count_xlsx_file_path = data_handler.get_current_daily_file_path() + "/" + file
            break
    time.sleep(1)
    kitchen_work_count_xlsx_wb = openpyxl.load_workbook(kitchen_work_count_xlsx_file_path)
    kitchen_work_count_xlsx_ws = kitchen_work_count_xlsx_wb.active

    kitchen_work_count_xlsx_ws['B1'] = int(morning_table_count) + int(evening_table_count)
    kitchen_work_count_xlsx_ws['B2'] = data_handler.get_server_date()
    kitchen_work_count_xlsx_wb.save(kitchen_work_count_xlsx_file_path)
    kitchen_work_count_xlsx_wb.close()


def daily_work_hour_count():
    daily_data = data_handler.get_daily_data()
    xlsx_file_path = ""
    for file in os.listdir(destination_path):
        if "日工时记录" in file:
            xlsx_file_path = destination_path + "/" + file
            break
    time.sleep(1)
    xlsx_wb = openpyxl.load_workbook(xlsx_file_path)
    xlsx_ws = xlsx_wb.active

    # count appetizer work hour
    appetizer_work_hour_morning = 0
    appetizer_work_hour_evening = 0
    for employee in daily_data['appetizer_work_hour']:
        if daily_data['appetizer_work_hour'][employee]['shift'] == 'morning':
            appetizer_work_hour_morning += float(daily_data['appetizer_work_hour'][employee]['worked_hour'])
        else:
            appetizer_work_hour_evening += float(daily_data['appetizer_work_hour'][employee]['worked_hour'])

    xlsx_ws['B3'] = appetizer_work_hour_morning
    xlsx_ws['D3'] = appetizer_work_hour_evening

    # count cleaner work hour
    total_cleaner_hour = 0
    for employee in daily_data['cleaner_work_hour']:
        total_cleaner_hour += float(daily_data['cleaner_work_hour'][employee])

    xlsx_ws['G5'] = total_cleaner_hour/2
    xlsx_ws['I5'] = total_cleaner_hour/2

    # count dish prepare work hour
    total_dish_prepare_hour = 0
    for item in daily_data['dish_prepare_work_hour']:
        if 'work_hour' in item and daily_data['dish_prepare_work_hour'][item] != "":
            total_dish_prepare_hour += float(daily_data['dish_prepare_work_hour'][item])

    xlsx_ws['B4'] = total_dish_prepare_hour/2
    xlsx_ws['D4'] = total_dish_prepare_hour/2

    # count dish runner work hour
    dish_runner_work_hour_morning = 0
    dish_runner_work_hour_evening = 0
    for employee in daily_data['dish_runner_work_hour']:
        dish_runner_work_hour_morning += float(daily_data['dish_runner_work_hour'][employee]['morning_worked_hour'])
        dish_runner_work_hour_evening += float(daily_data['dish_runner_work_hour'][employee]['evening_worked_hour'])

    xlsx_ws['B5'] = dish_runner_work_hour_morning
    xlsx_ws['D5'] = dish_runner_work_hour_evening

    # count host work hour
    host_work_hour_morning = 0
    host_work_hour_evening = 0
    for employee in daily_data['host_work_hour']:
        if daily_data['host_work_hour'][employee]['shift'] == 'morning':
            host_work_hour_morning += float(daily_data['host_work_hour'][employee]['worked_hour'])
        else:
            host_work_hour_evening += float(daily_data['host_work_hour'][employee]['worked_hour'])

    xlsx_ws['G4'] = host_work_hour_morning
    xlsx_ws['I4'] = host_work_hour_evening

    # count noodle dancer work hour
    total_noodle_dancer_hour = 0
    for employee in daily_data['noodle_dance_work_hour']:
        total_noodle_dancer_hour += float(daily_data['noodle_dance_work_hour'][employee])

    xlsx_ws['B7'] = total_noodle_dancer_hour/2
    xlsx_ws['D7'] = total_noodle_dancer_hour/2

    # count server work hour
    morning_server_hour = 0
    evening_server_hour = 0
    for group in daily_data['server_group_info']:
        current_group = daily_data['server_group_info'][group]
        group_total_hour = 0
        for item in current_group:
            if 'work_hour' in item and 'transfer' not in item and current_group[item] != "":
                group_total_hour += float(current_group[item])
        if current_group['shift'] == "morning":
            morning_server_hour += group_total_hour
        else:
            evening_server_hour += group_total_hour

    xlsx_ws['G3'] = morning_server_hour
    xlsx_ws['I3'] = evening_server_hour

    xlsx_wb.save(xlsx_file_path)
    xlsx_wb.close()


def daily_close_and_get_file():
    init_files()
    # get_haidilao_pos_data() # 因为服务器没有vpn所以没有办法从海底捞系统里拿资料
    busser_work_count()
    kitchen_work_count()
    daily_tips_count()
    daily_take_out_bonus_count()
    daily_work_hour_count()
    shutil.make_archive(destination_path, 'zip', destination_path)
    data_handler.set_date_today(str(datetime.date.today()).replace("-", ""))
    return destination_path + '.zip'
