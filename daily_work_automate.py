"""
此文件为单独计算HDL日常清账使用
制作人：Hongming Wang
"""
import csv
import datetime
import json
import pathlib
import re
import shutil
import time
import os, sys
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl
import sms_handler


# setups
YMD = None
YEAR = None
MONTH = None
DAY = None
NextDay_Date = None
NextDay_year = None
NextDay_month = None
NextDay_day = None

destination_path = ""
website_url = "http://127.0.0.1:8894/"

# accounts
HAIDILAO_USERNAME = ""
HAIDILAO_PASSWORD = ""

IOT_USERNAME = ""
IOT_PASSWORD = ""
IOT_MERCHANT_ID = ""

UBER_USERNAME = "@ubereats.com"
UBER_PASSWORD = ""
UBER_PIN = ""
UBER_File_include_string = ""

SNAPPY_USERNAME = ""
SNAPPY_PASSWORD = ""
SNAPPY_STORE_ID = ""

FANTUAN_USERNAME = ""
FANTUAN_PASSWORD = ""


def pos_haidilao_daily():
    driver = init_web_driver()
    print("进入海底捞网站")
    driver.get('https://pos.haidilao.com/')
    driver.find_element(by=By.CLASS_NAME, value="el-form-item__content").find_element(by=By.CLASS_NAME,
                                                                                      value="el-input__inner").send_keys(
        HAIDILAO_USERNAME)
    driver.find_element(by=By.CLASS_NAME, value="el-form-item.password.is-required").find_element(by=By.CLASS_NAME,
                                                                                                  value="el-input__inner").send_keys(
        HAIDILAO_PASSWORD)
    driver.find_element(by=By.CLASS_NAME, value="loginButton").click()

    try:
        main = WebDriverWait(driver, 10).until(
            expected_conditions.presence_of_element_located((By.CLASS_NAME, "content-col.el-col.el-col-6")))
        driver.find_element(by=By.CLASS_NAME, value="content-col.el-col.el-col-6").click()
    except Exception as e:
        print(e)
        driver.quit()
    child_child_side_buttons = None
    print("进入清账日结表")
    try:
        WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.CLASS_NAME, "el-submenu")))
        side_buttons = driver.find_elements(by=By.CLASS_NAME, value="el-submenu")
        right_side_button = None
        for button in side_buttons:
            try:
                button.find_element(By.CLASS_NAME, "icon.iconfont.icon-reportmgr")
                button.click()
                right_side_button = button
                break
            except Exception as e:
                pass
        child_side_buttons = right_side_button.find_elements(by=By.CLASS_NAME, value="el-menu-item-group")
        time.sleep(1)
        right_child_side_button = None
        for button in child_side_buttons:
            try:
                menu_name = button.find_element(By.CLASS_NAME, "el-submenu__title").text
                if menu_name == "营业报表":
                    button.click()
                    right_child_side_button = button
                    break
            except:
                pass
        time.sleep(1)
        child_child_side_buttons = right_child_side_button.find_elements(by=By.CLASS_NAME, value="el-menu-item-group")
        for button in child_child_side_buttons:
            try:
                menu_name = button.find_element(By.CLASS_NAME, "el-menu-item").text
                if menu_name == "清帐日结表":
                    button.click()
                    break
            except:
                pass
    except Exception as e:
        print(e)
        driver.quit()
    time.sleep(1)
    print("操作清账日结表")
    try:
        main = WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located(
            (By.CLASS_NAME, "el-date-editor.el-input.el-date-editor--date")))
        driver.find_element(By.CLASS_NAME, "el-date-editor.el-input.el-date-editor--date").find_element(By.CLASS_NAME,
                                                                                                        "el-input__inner").clear()
        driver.find_element(By.CLASS_NAME, "el-date-editor.el-input.el-date-editor--date").find_element(By.CLASS_NAME,
                                                                                                        "el-input__inner").send_keys(
            f"{YEAR}-{MONTH}-{DAY}")
        primary_buttons = driver.find_elements(By.CLASS_NAME, "el-button.el-button--primary")
        for button in primary_buttons:
            if button.text == "重新日结":
                button.click()
                break
        time.sleep(1)
        driver.find_element(By.CLASS_NAME, "el-button.el-button--default.el-button--primary ").click()
        time.sleep(15)
        for button in primary_buttons:
            if button.text == "查询":
                button.click()
                break
        time.sleep(5)
        data_columns = driver.find_element(By.CLASS_NAME, "daily-table").find_elements(By.CSS_SELECTOR, "tr")
        is_operating_columns = False
        pos_cash_amount = 0
        daily_close_person_name = "Hongming Wang"
        daily_close_person_number = "7788611008"
        for column in data_columns:  # 填写空格子
            if is_operating_columns is True:
                if "现金" in column.text:
                    for div in column.find_elements(By.XPATH, ".//*"):
                        try:
                            float(div.text)
                            pos_cash_amount = float(div.text)
                        except:
                            continue
                elif "大钞" in column.text:
                    column.find_element(By.CLASS_NAME, "el-input__inner").clear()
                    column.find_element(By.CLASS_NAME, "el-input__inner").send_keys(pos_cash_amount)
                elif "结算合计" in column.text:
                    pass
                else:
                    for div in column.find_elements(By.XPATH, ".//*"):
                        try:
                            float(div.text)
                            column.find_element(By.CLASS_NAME, "el-input__inner").clear()
                            column.find_element(By.CLASS_NAME, "el-input__inner").send_keys(div.text)
                            break
                        except:
                            continue
            if "营业收入合计" in column.text:
                if column.text.split(" ")[1] == "0":
                    print("!!!海底捞无法清账!!!")
                    return
            if "结算方式" in column.text:
                is_operating_columns = True
            if "非营业收入及小费" in column.text:
                is_operating_columns = False
            if "收银员" in column.text:
                input_list = column.find_elements(By.CLASS_NAME, "el-input__inner")
                for input_block in input_list:
                    input_block.clear()
                    input_block.send_keys(daily_close_person_name)
            if "收银员手机号码" in column.text:
                input_list = column.find_elements(By.CLASS_NAME, "el-input__inner")
                for input_block in input_list:
                    input_block.clear()
                    input_block.send_keys(daily_close_person_number)
        primary_buttons = driver.find_elements(By.CLASS_NAME, "el-button.el-button--primary")
        for button in primary_buttons:
            if button.text == "保存":
                button.click()
        time.sleep(1)
        driver.find_element(By.CLASS_NAME, "el-button.el-button--default.el-button--primary ").click()
        time.sleep(10)
        print("从海底捞网页下载日结表")
        for button in primary_buttons:
            if button.text == "导出":
                button.click()
                break
        time.sleep(10)
        for file in os.listdir(os.getcwd() + "\\downloaded_files"):
            if "日结单报表" in file:
                shutil.copy(os.getcwd() + "\\downloaded_files\\" + file, destination_path + f"\\{YMD}收账明细\\" + file)
                break
        time.sleep(10)
        print("处理海底捞下载文件")
        close_book_path = None  # 收账本模板
        daily_close_book_path = None  # 海底捞pos下载日结报表单
        for file in os.listdir(destination_path + f"\\{YMD}收账明细"):
            if "收账本" in file:
                close_book_path = destination_path + f"\\{YMD}收账明细\\" + file
            if "日结单报表" in file:
                daily_close_book_path = destination_path + f"\\{YMD}收账明细\\" + file
        close_book_wb = openpyxl.load_workbook(close_book_path)
        daily_close_book_wb = openpyxl.load_workbook(daily_close_book_path)
        daily_close_book_workbook_ws = close_book_wb.active
        daily_close_data_workbook_ws = daily_close_book_wb.active
        print("从日结表单输入新的数据")
        for col in range(1, 35):
            if daily_close_data_workbook_ws["D" + str(col)].value == "现金":
                daily_close_book_workbook_ws["G7"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
            elif daily_close_data_workbook_ws["D" + str(col)].value == "信用卡":
                daily_close_book_workbook_ws["G8"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
            elif daily_close_data_workbook_ws["D" + str(col)].value == "饭团":
                daily_close_book_workbook_ws["G10"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
            elif daily_close_data_workbook_ws["D" + str(col)].value == "线下支付宝":
                if daily_close_book_workbook_ws["G9"].value is not None:
                    daily_close_book_workbook_ws["G9"] = float(daily_close_data_workbook_ws["F" + str(col)].value) + \
                                                         float(daily_close_book_workbook_ws["G9"].value)
                else:
                    daily_close_book_workbook_ws["G9"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
            elif daily_close_data_workbook_ws["D" + str(col)].value == "线下微信":
                if daily_close_book_workbook_ws["G9"].value is not None:
                    daily_close_book_workbook_ws["G9"] = float(daily_close_data_workbook_ws["F" + str(col)].value) + \
                                                         float(daily_close_book_workbook_ws["G9"].value)
                else:
                    daily_close_book_workbook_ws["G9"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
            elif daily_close_data_workbook_ws["D" + str(col)].value == "Snappy":
                daily_close_book_workbook_ws["G16"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
            elif daily_close_data_workbook_ws["D" + str(col)].value == "uber":
                daily_close_book_workbook_ws["G14"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
            elif daily_close_data_workbook_ws["D" + str(col)].value == "Vanpeople":
                daily_close_book_workbook_ws["G19"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
            elif daily_close_data_workbook_ws["D" + str(col)].value == "Tapin":
                daily_close_book_workbook_ws["G18"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
            elif daily_close_data_workbook_ws["D" + str(col)].value == "Panda":
                daily_close_book_workbook_ws["G17"] = float(daily_close_data_workbook_ws["F" + str(col)].value)
            close_book_wb.save(close_book_path)
        for button in child_child_side_buttons:
            try:
                menu_name = button.find_element(By.CLASS_NAME, "el-menu-item").text
                if menu_name == "日营业汇总表":
                    button.click()
                    break
            except:
                pass
        print("输入日营业汇总表销售数据")
        main = WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located(
            (By.CLASS_NAME, "el-date-editor.el-input.el-input--small.el-date-editor--daterange")))
        driver.find_element(by=By.CLASS_NAME,
                            value="el-date-editor.el-input.el-input--small.el-date-editor--daterange").find_element(
            by=By.CLASS_NAME, value="el-input__inner").clear()
        driver.find_element(by=By.CLASS_NAME,
                            value="el-date-editor.el-input.el-input--small.el-date-editor--daterange").find_element(
            by=By.CLASS_NAME, value="el-input__inner").send_keys(f'{YEAR}-{MONTH}-{DAY},{YEAR}-{MONTH}-{DAY}')
        time.sleep(2)
        driver.find_element(by=By.CLASS_NAME, value="el-button.el-button--primary.el-button--small").click()
        time.sleep(2)
        dine_in_money = driver.find_element(by=By.CLASS_NAME, value="el-table__body-wrapper").find_element(
            by=By.CLASS_NAME, value="el-table_1_column_6").text
        takeout_money = driver.find_element(by=By.CLASS_NAME, value="el-table__body-wrapper").find_element(
            by=By.CLASS_NAME, value="el-table_1_column_7").text
        retail_money = driver.find_element(by=By.CLASS_NAME, value="el-table__body-wrapper").find_element(
            by=By.CLASS_NAME, value="el-table_1_column_8").text
        daily_close_book_workbook_ws["B2"] = dine_in_money
        daily_close_book_workbook_ws["B3"] = takeout_money
        daily_close_book_workbook_ws["B1"] = retail_money
        close_book_wb.save(close_book_path)
        close_book_wb.close()
        daily_close_book_wb.close()
        print("海底捞后台处理完成")
        driver.quit()
    except Exception as e:
        print(e)
        driver.quit()


def iot_daily():
    driver = init_web_driver()
    print("登入IOT账号")
    driver.get('https://admin.iotpaycloud.com/login')
    driver.find_element(By.ID, "username").send_keys(IOT_USERNAME)
    driver.find_element(By.ID, "password").send_keys(IOT_PASSWORD)
    driver.find_element(By.CLASS_NAME, "btn.btn-block.btn-primary").click()
    print("下载IOT日常报表")
    driver.get('https://admin.iotpaycloud.com/orders/daily')
    """ 
    # 发现可以直接进入网页地址不需要一个一个按
    side_menus = driver.find_elements(By.CLASS_NAME, "menu-wrapper")
    order_list_menu = None
    for menu in side_menus:
        if "订单列表" in menu.text or "Order List" in menu.text:
            order_list_menu = menu
            menu.click()
            break
    order_list_childs = order_list_menu.find_element(By.CLASS_NAME, "el-menu.el-menu--inline").find_elements(By.CLASS_NAME, "el-menu-item")
    for order_list_child in order_list_childs:
        if "Daily" in order_list_child.text or "每日" in order_list_child.text:
            order_list_child.click()
            break
    """
    WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.CLASS_NAME, "el-table__row")))
    daily_data_rows = driver.find_element(By.CLASS_NAME, "card-body").find_elements(By.CLASS_NAME, "el-table__row")
    for row in daily_data_rows:
        if f"{YEAR}-{MONTH}-{DAY}" in row.text:
            row.find_element(By.CLASS_NAME, "el-button.icon-btn.el-button--success.el-button--mini").click()
    time.sleep(5)
    for file in os.listdir(os.getcwd() + "\\downloaded_files"):
        if IOT_MERCHANT_ID in file:
            shutil.copy(os.getcwd() + "\\downloaded_files\\" + file, destination_path + f"\\{YMD}收账明细\\IOT.csv")
            break
    time.sleep(5)
    print("处理IOT日常报表")
    close_book_path = None  # 收账本模板
    iot_daily_file_path = None  # IOT下载表单
    for file in os.listdir(destination_path + f"\\{YMD}收账明细"):
        if "收账本" in file:
            close_book_path = destination_path + f"\\{YMD}收账明细\\" + file
        if "IOT" in file:
            iot_daily_file_path = destination_path + f"\\{YMD}收账明细\\" + file
    close_book_wb = openpyxl.load_workbook(close_book_path)
    daily_close_book_workbook_ws = close_book_wb.active
    iot_daily_file = open(iot_daily_file_path, "r")
    total_iot_money = 0
    total_iot_tips = 0
    for line in iot_daily_file:
        items = line.split(",")
        try:
            if items[20].split(" ")[1].split(":")[0].isdigit() and int(items[20].split(" ")[1].split(":")[0]) > 9:
                total_iot_money += float(items[15])
                total_iot_tips += float(items[16])
        except:
            pass
    iot_daily_file.close()
    driver.get('https://admin.iotpaycloud.com/orders/today')
    WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.CLASS_NAME, "el-table__row")))
    daily_data_rows = driver.find_element(By.CLASS_NAME, "el-table__body").find_elements(By.CLASS_NAME, "el-table__row")
    for item in daily_data_rows:
        item.find_element(By.CLASS_NAME, "el-table__expand-icon").click()
    daily_data_expanded_cells = driver.find_elements(By.CLASS_NAME, "el-table__expanded-cell")
    for cell in daily_data_expanded_cells:
        cell_rows = cell.find_elements(By.CLASS_NAME, "el-form-item.el-form-item--small")
        for row in cell_rows:
            if "Create Time" in row.text or "创建时间" in row.text:
                try:
                    if int(row.text.split(" ")[1].split(":")[0]) < 9:
                        for row in cell_rows:
                            label = row.find_element(By.CLASS_NAME, "el-form-item__label").text.strip()
                            if "Amount" == label or "金额" == label:
                                total_iot_money += float(row.text.split("$")[1])
                        for row in cell_rows:
                            label = row.find_element(By.CLASS_NAME, "el-form-item__label").text.strip()
                            if "Tip" == label or "小费" == label:
                                total_iot_tips += float(row.text.split("$")[1])
                        break
                except:
                    pass
    daily_close_book_workbook_ws["C9"] = f"={total_iot_money}-D9"
    daily_close_book_workbook_ws["D9"] = total_iot_tips
    close_book_wb.save(close_book_path)
    close_book_wb.close()
    driver.quit()
    print("完成IOT清账")


def uber_eat_daily():
    print("登入Uber eats账号")
    driver = init_web_driver()
    driver.get('https://merchants.ubereats.com/manager/home/6caef017-314a-453a-b658-e5d6957532a7')
    # 账号部分
    time.sleep(2)
    driver.find_element(By.CLASS_NAME, "text-input").clear()
    driver.find_element(By.CLASS_NAME, "text-input").send_keys(UBER_USERNAME)
    driver.find_element(By.CLASS_NAME, "btn.btn--arrow.btn--full").click()
    time.sleep(5)
    # 密码部分
    driver.find_element(By.XPATH, '//*[@id="password"]').clear()
    driver.find_element(By.XPATH, '//*[@id="password"]').send_keys(UBER_PASSWORD)
    driver.find_element(By.CLASS_NAME, "btn.btn--arrow.btn--full").click()
    time.sleep(5)
    # pin code部分
    try:
        WebDriverWait(driver, 5).until(expected_conditions.presence_of_element_located((By.CLASS_NAME, "c7.dr.as.au.at.av.aw.ay.ax.az.b0.c6.ds.dt.du.dv.dw.dx.dy.dz.bf.bg.dg.dh.cx.e0.e1.e2")))
        driver.find_element(By.CLASS_NAME, "c7.dr.as.au.at.av.aw.ay.ax.az.b0.c6.ds.dt.du.dv.dw.dx.dy.dz.bf.bg.dg.dh.cx.e0.e1.e2").send_keys(UBER_PIN)
        driver.find_element(By.CLASS_NAME, "ao.ap.aq.ar.as.at.au.av.aw.ax.ay.az.b0.b1.b2.b3.b4.b5.b6.b7.b8.b9.ba.bb.eb.bd.be.bf.bg.bh.bi.de.df.dc.dd.bn.bo.ec.ed.br.bs.bt.bu.ee.ef").click()
        time.sleep(2)
    except:
        pass
    WebDriverWait(driver, 2).until(
        expected_conditions.presence_of_element_located((By.CLASS_NAME, "bf.ca.bh.cb.co.cp.cf.cq")))
    print("生成uber报表")
    driver.get('https://merchants.ubereats.com/manager/reports?restaurantUUID=6caef017-314a-453a-b658-e5d6957532a7')
    # 点击索取报告
    WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.XPATH, "//*[@id='wrapper']/div[1]/div[2]/div[2]/div[2]/div/div[1]/div/div[2]/div/button")))
    driver.find_element(By.XPATH, "//*[@id='wrapper']/div[1]/div[2]/div[2]/div[2]/div/div[1]/div/div[2]/div/button").click()
    # 点击报告内框-选择报表类型
    WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/div[1]/div[2]/div[2]/div[2]/div/div[3]/div[2]/div')))
    driver.find_element(By.XPATH, '//*[@id="wrapper"]/div[1]/div[2]/div[2]/div[2]/div/div[3]/div[2]/div').click()
    time.sleep(1)
    # 选择报表类型
    report_types = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div[2]/div/div/div/div/div/ul/li[1]').click()
    '''
    for type in report_types:
        if "收入详细信息" in type.text or "Payment Details" in type.text:
            type.click()
            break
    '''
    time.sleep(1)
    # 输入报表起始结束日期
    driver.find_element(By.XPATH, '//*[@id="wrapper"]/div[1]/div[2]/div[2]/div[2]/div/div[4]/div[2]/div/div').click()
    driver.find_element(By.XPATH, '//*[@id="wrapper"]/div[1]/div[2]/div[2]/div[2]/div/div[4]/div[2]/div/div/input').send_keys(f"{YMD}{NextDay_Date}")
    # 点击索取报表
    WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.XPATH, "//*[@id='wrapper']/div[1]/div[2]/div[2]/div[2]/div/div[1]/button[2]")))
    driver.find_element(By.XPATH, "//*[@id='wrapper']/div[1]/div[2]/div[2]/div[2]/div/div[1]/button[2]").click()
    print("下载uber报表")
    time.sleep(30)
    driver.find_element(By.XPATH, "//*[@id='wrapper']/div[1]/div[2]/div[2]/div[2]/div/div[4]/div[2]/div[1]/div/div[4]/div").click()
    time.sleep(5)
    for file in os.listdir(os.getcwd() + "\\downloaded_files"):
        if UBER_File_include_string in file:
            shutil.copy(os.getcwd() + "\\downloaded_files\\" + file, destination_path + f"\\{YMD}收账明细\\Uber.csv")
            break
    time.sleep(5)
    close_book_path = None  # 收账本模板
    uber_daily_file_path = None  # Uber下载表单
    for file in os.listdir(destination_path + f"\\{YMD}收账明细"):
        if "收账本" in file:
            close_book_path = destination_path + f"\\{YMD}收账明细\\" + file
        if "Uber" in file:
            uber_daily_file_path = destination_path + f"\\{YMD}收账明细\\" + file
    time.sleep(5)
    print("处理uber报表数据")
    close_book_wb = openpyxl.load_workbook(close_book_path)
    daily_close_book_workbook_ws = close_book_wb.active
    uber_daily_file = open(uber_daily_file_path, "r", encoding="utf8")
    uber_total_amount = 0
    for line in uber_daily_file:
        items = line.split(",")
        try:
            if int(items[4].split("-")[2]) == int(DAY) and int(items[5].split(":")[0]) >= 9 or int(items[4].split("-")[2]) == int(NextDay_day) and int(items[5].split(":")[0]) <= 9:
                uber_total_amount += float(items[22])
        except:
            pass
    uber_daily_file.close()
    daily_close_book_workbook_ws["C14"] = float(uber_total_amount)
    close_book_wb.save(close_book_path)
    close_book_wb.close()
    driver.quit()
    print("Uber eat日结完成")


def snappy_daily():
    driver = init_web_driver()
    print("登入Snappy账号")
    driver.get('http://gosnappy.io/bc2/login')
    driver.find_element(By.XPATH, '//*[@id="username"]').send_keys(SNAPPY_USERNAME)
    driver.find_element(By.XPATH, '//*[@id="password"]').send_keys(SNAPPY_PASSWORD)
    driver.find_element(By.XPATH, '//*[@id="storeid"]').send_keys(SNAPPY_STORE_ID)
    driver.find_element(By.CLASS_NAME, 'btn.btn-primary.mat-raised-button.mat-primary.ng-star-inserted').click()
    time.sleep(5)
    driver.get('http://gosnappy.io/bc2/transactions')
    time.sleep(3)
    print("读取Snappy资料")
    driver.execute_script(f'document.getElementById("mat-input-0").removeAttribute("readonly")')
    driver.execute_script(f'document.getElementById("mat-input-0").setAttribute("class","")')
    driver.execute_script(f'document.getElementById("mat-input-0").value = null')
    driver.find_element(By.ID, 'mat-input-0').send_keys(f"{int(MONTH)}/{int(DAY)}/{int(YEAR)} - {int(NextDay_month)}/{int(NextDay_day)}/{int(NextDay_year)}")
    time.sleep(3)
    driver.find_element(By.XPATH, '/html/body/app-root/div/mat-drawer-container/mat-drawer-content/app-transactions/div/div[2]/div/button[1]').click()
    time.sleep(3)
    print("下载Snappy报表")
    driver.find_element(By.XPATH, '/html/body/app-root/div/mat-drawer-container/mat-drawer-content/app-transactions/div/div[2]/div/button[2]').click()
    time.sleep(5)
    for file in os.listdir(os.getcwd() + "\\downloaded_files"):
        if SNAPPY_STORE_ID in file:
            shutil.copy(os.getcwd() + "\\downloaded_files\\" + file, destination_path + f"\\{YMD}收账明细\\Snappy.csv")
            break
    time.sleep(5)
    print("处理Snappy报表数据")
    close_book_path = None  # 收账本模板
    snappy_daily_file_path = None  # Snappy下载表单
    for file in os.listdir(destination_path + f"\\{YMD}收账明细"):
        if "收账本" in file:
            close_book_path = destination_path + f"\\{YMD}收账明细\\" + file
        if "Snappy" in file:
            snappy_daily_file_path = destination_path + f"\\{YMD}收账明细\\" + file
    time.sleep(5)
    snappy_total_amount = 0
    close_book_wb = openpyxl.load_workbook(close_book_path)
    daily_close_book_workbook_ws = close_book_wb.active
    with open(snappy_daily_file_path, "r") as csvfile:
        reader = csv.reader(csvfile)
        for line in reader:
            items = line
            try:
                if (items[0].split("T")[0].split("-")[2] == DAY and int(items[0].split("T")[1].split(":")[0]) > 10) or \
                        (items[0].split("T")[0].split("-")[2] == NextDay_day and int(
                            items[0].split("T")[1].split(":")[0]) < 10):
                    print(float(items[8]))
                    snappy_total_amount += float(items[8])
            except Exception as e:
                print(e)
    daily_close_book_workbook_ws["C16"] = float(snappy_total_amount)
    close_book_wb.save(close_book_path)
    close_book_wb.close()
    driver.quit()
    print("Snappy日结完成")


def fantuan_daily():
    ft_amounts = 0
    ft_fee = 0
    driver = init_web_driver()
    print("登入饭团账号")
    driver.get('https://pos.fantuan.ca/')
    driver.find_element(By.XPATH, '//*[@placeholder="用户名"]').send_keys(FANTUAN_USERNAME)
    driver.find_element(By.XPATH, '//*[@placeholder="密码"]').send_keys(FANTUAN_PASSWORD)
    driver.find_element(By.XPATH, '//*[@type="submit"]').click()
    time.sleep(5)
    try:
        driver.find_element(By.CLASS_NAME, 'send_smsCode_btn').click()
        time.sleep(30)
        driver.find_element(By.XPATH, '//*[@type="tel"]').send_keys(
            re.findall("\d{6}", sms_handler.get_most_recent_message())[0])
        driver.find_element(By.CLASS_NAME, 'verify_smsCode_btn').click()
        time.sleep(5)
    except:
        pass
    print("读取饭团后台数据")
    # 读取当日单据数据
    driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div[1]/div/div/div[2]/div/div/div[2]').click()
    time.sleep(5)
    left_right_buttons = driver.find_elements(By.XPATH, '//*[@class="__1ZaR-"]')
    left_right_buttons[1].click()
    time.sleep(5)
    try:
        orders = driver.find_elements(By.XPATH,
                                      '//*[@id="root"]/div/div[1]/div[2]/div[1]/div/div/div[4]/div[1]/div/div[2]/div')
        for item in orders:
            split_items = item.text.split()
            if len(split_items) > 1:
                if int(split_items[3].split(":")[0]) > 10:
                    item.click()
                    time.sleep(0.5)
                    WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.CLASS_NAME, '__33MCY.__2QIVY')))
                    amounts = driver.find_elements(By.CLASS_NAME, '__33MCY.__2QIVY')
                    if amounts[2].text.split()[0] == "Fantuan Transfer" or amounts[2].text.split()[0] == "饭团转账":
                        fantuan_transfer = amounts[2].text.split()[1].replace("$", "")
                        fantuan_fee = amounts[1].text.split()[1].replace("-$", "")
                        ft_amounts += float(fantuan_transfer)
                        ft_fee += float(fantuan_fee)
    except:
        pass
    # 读取凌晨单据数据
    left_right_buttons = driver.find_elements(By.XPATH, '//*[@class="__1ZaR-"]')
    left_right_buttons[2].click()
    time.sleep(5)
    try:
        orders = driver.find_elements(By.XPATH,
                                      '//*[@id="root"]/div/div[1]/div[2]/div[1]/div/div/div[4]/div[1]/div/div[2]/div')
        for item in orders:
            split_items = item.text.split()
            if len(split_items) > 1:
                if int(split_items[3].split(":")[0]) < 10:
                    item.click()
                    time.sleep(0.5)
                    WebDriverWait(driver, 10).until(
                        expected_conditions.presence_of_element_located((By.CLASS_NAME, '__33MCY.__2QIVY')))
                    amounts = driver.find_elements(By.CLASS_NAME, '__33MCY.__2QIVY')
                    if amounts[2].text.split()[0] == "Fantuan Transfer" or amounts[2].text.split()[0] == "饭团转账":
                        fantuan_transfer = amounts[2].text.split()[1].replace("$", "")
                        fantuan_fee = amounts[1].text.split()[1].replace("-$", "")
                        ft_amounts += float(fantuan_transfer)
                        ft_fee += float(fantuan_fee)
    except:
        pass
    print("将饭团后台数据导入收账本")
    close_book_path = None  # 收账本模板
    for file in os.listdir(destination_path + f"\\{YMD}收账明细"):
        if "收账本" in file:
            close_book_path = destination_path + f"\\{YMD}收账明细\\" + file
    time.sleep(5)
    close_book_wb = openpyxl.load_workbook(close_book_path)
    daily_close_book_workbook_ws = close_book_wb.active
    daily_close_book_workbook_ws["C10"] = float(ft_amounts)
    daily_close_book_workbook_ws["E10"] = float(ft_fee)
    close_book_wb.save(close_book_path)
    close_book_wb.close()
    driver.quit()
    print("饭团日结完成")


def busser_work_count():
    employee_data = get_data_from_employee_data_json_file()
    daily_data = get_data_from_daily_data_json_file()
    busser_work_count_xlsx_file_path = ""
    for file in os.listdir(destination_path):
        if "传配组打卡计件" in file:
            busser_work_count_xlsx_file_path = destination_path + "\\" + file
            break
    time.sleep(5)
    busser_work_count_xlsx_wb = openpyxl.load_workbook(busser_work_count_xlsx_file_path)
    busser_work_count_xlsx_ws = busser_work_count_xlsx_wb.active

    for row in busser_work_count_xlsx_ws['A13':'E49']:
        for cell in row:
            cell.value = None
    for row in busser_work_count_xlsx_ws['B2':'B7']:
        for cell in row:
            cell.value = None

    start_row = 13
    for person in daily_data[date]["dish_runner_work_hour"]:
        busser_work_count_xlsx_ws[f"A{start_row}"] = str(person).split(" ")[0]
        busser_work_count_xlsx_ws[f"B{start_row}"] = str(person).split(" ", 1)[1]
        try:
            busser_work_count_xlsx_ws[f"C{start_row}"] = float(employee_data[str(person).split(" ")[0]]["dish_runner_work_count_ratio"])
        except Exception as e:
            busser_work_count_xlsx_ws[f"C{start_row}"] = 0
        busser_work_count_xlsx_ws[f"D{start_row}"] = float(daily_data[date]["dish_runner_work_hour"][person]["morning_worked_hour"])
        busser_work_count_xlsx_ws[f"E{start_row}"] = float(daily_data[date]["dish_runner_work_hour"][person]["evening_worked_hour"])
        start_row += 1
    busser_work_count_xlsx_wb.save(busser_work_count_xlsx_file_path)
    busser_work_count_xlsx_wb.close()


def kitchen_work_count():
    employee_data = get_data_from_employee_data_json_file()
    daily_data = get_data_from_daily_data_json_file()
    kitchen_work_count_xlsx_file_path = ""
    for file in os.listdir(destination_path):
        if "厨房计件" in file:
            kitchen_work_count_xlsx_file_path = destination_path + "\\" + file
            break
    time.sleep(5)
    kitchen_work_count_xlsx_wb = openpyxl.load_workbook(kitchen_work_count_xlsx_file_path)
    kitchen_work_count_xlsx_ws = kitchen_work_count_xlsx_wb.active

    for row in kitchen_work_count_xlsx_ws['B1':'B2']:
        for cell in row:
            cell.value = None
    for row in kitchen_work_count_xlsx_ws['A6':'C9']:
        for cell in row:
            cell.value = None
    for row in kitchen_work_count_xlsx_ws['A14':'C17']:
        for cell in row:
            cell.value = None
    for row in kitchen_work_count_xlsx_ws['A22':'C25']:
        for cell in row:
            cell.value = None
    for row in kitchen_work_count_xlsx_ws['A30':'C33']:
        for cell in row:
            cell.value = None
    for row in kitchen_work_count_xlsx_ws['A38':'C41']:
        for cell in row:
            cell.value = None
    for row in kitchen_work_count_xlsx_ws['A46':'C53']:
        for cell in row:
            cell.value = None

    # 日期
    kitchen_work_count_xlsx_ws["B2"] = f"{YEAR}-{MONTH}-{DAY}"

    #上菜口
    row_number = 6
    if daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window1"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window1"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window1"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window1_work_hour"])
    row_number = 7
    if daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window2"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window2"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window2"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window2_work_hour"])
    row_number = 8
    if daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window3"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window3"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window3"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window3_work_hour"])
    row_number = 9
    if daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window4"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window4"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window4"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["dish_prepare_window4_work_hour"])

    #羊肉间
    row_number = 14
    if daily_data[date]["dish_prepare_work_hour"]["lamb_room1"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["lamb_room1"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["lamb_room1"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["lamb_room1_work_hour"])
    row_number = 15
    if daily_data[date]["dish_prepare_work_hour"]["lamb_room2"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["lamb_room2"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["lamb_room2"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["lamb_room2_work_hour"])
    row_number = 16
    if daily_data[date]["dish_prepare_work_hour"]["lamb_room3"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["lamb_room3"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["lamb_room3"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["lamb_room3_work_hour"])
    row_number = 17
    if daily_data[date]["dish_prepare_work_hour"]["lamb_room4"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["lamb_room4"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["lamb_room4"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["lamb_room4_work_hour"])

    # 海鲜
    row_number = 22
    if daily_data[date]["dish_prepare_work_hour"]["sea_food1"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["sea_food1"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["sea_food1"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["sea_food1_work_hour"])
    row_number = 23
    if daily_data[date]["dish_prepare_work_hour"]["sea_food2"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["sea_food2"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["sea_food2"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["sea_food2_work_hour"])
    row_number = 24
    if daily_data[date]["dish_prepare_work_hour"]["sea_food3"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["sea_food3"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["sea_food3"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["sea_food3_work_hour"])
    row_number = 25
    if daily_data[date]["dish_prepare_work_hour"]["sea_food4"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["sea_food4"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["sea_food4"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["sea_food4_work_hour"])

    # 豆菌
    row_number = 30
    if daily_data[date]["dish_prepare_work_hour"]["fungus_room1"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["fungus_room1"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["fungus_room1"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["fungus_room1_work_hour"])
    row_number = 31
    if daily_data[date]["dish_prepare_work_hour"]["fungus_room2"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["fungus_room2"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["fungus_room2"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["fungus_room2_work_hour"])
    row_number = 32
    if daily_data[date]["dish_prepare_work_hour"]["fungus_room3"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["fungus_room3"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["fungus_room3"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["fungus_room3_work_hour"])
    row_number = 33
    if daily_data[date]["dish_prepare_work_hour"]["fungus_room4"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["fungus_room4"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["fungus_room4"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["fungus_room4_work_hour"])

    # 粗加工
    row_number = 38
    if daily_data[date]["dish_prepare_work_hour"]["rough_process1"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["rough_process1"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["rough_process1"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["rough_process1_work_hour"])
    row_number = 39
    if daily_data[date]["dish_prepare_work_hour"]["rough_process2"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["rough_process2"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["rough_process2"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["rough_process2_work_hour"])
    row_number = 40
    if daily_data[date]["dish_prepare_work_hour"]["rough_process3"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["rough_process3"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["rough_process3"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["rough_process3_work_hour"])
    row_number = 41
    if daily_data[date]["dish_prepare_work_hour"]["rough_process4"] != "":
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["rough_process4"]).split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = \
            str(daily_data[date]["dish_prepare_work_hour"]["rough_process4"]).split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            float(daily_data[date]["dish_prepare_work_hour"]["rough_process4_work_hour"])

    #小吃房
    row_number = 46
    for person in daily_data[date]["appetizer_work_hour"]:
        kitchen_work_count_xlsx_ws[f"A{row_number}"] = person.split(" ")[0]
        kitchen_work_count_xlsx_ws[f"B{row_number}"] = person.split(" ", 1)[1]
        kitchen_work_count_xlsx_ws[f"C{row_number}"] = \
            str(daily_data[date]["appetizer_work_hour"][person])
        row_number += 1

    kitchen_work_count_xlsx_wb.save(kitchen_work_count_xlsx_file_path)
    kitchen_work_count_xlsx_wb.close()


def daily_tips_count():
    employee_data = get_data_from_employee_data_json_file()
    daily_data = get_data_from_daily_data_json_file()
    tips_count_xlsx_file_path = ""
    for file in os.listdir(destination_path):
        if "小费计算表格" in file:
            tips_count_xlsx_file_path = destination_path + "\\" + file
            break
    time.sleep(5)
    tips_count_xlsx_wb = openpyxl.load_workbook(tips_count_xlsx_file_path)
    tips_count_xlsx_ws = tips_count_xlsx_wb.active

    # 清除之前数据
    for row in tips_count_xlsx_ws['A4':'C103']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['F4':'G103']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['A108':'C124']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['E108':'E124']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['A129':'C137']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['E129':'E137']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['A142':'B151']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['A156':'C172']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['E156':'E172']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['A177':'C197']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['E177':'E197']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['A202':'C209']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['E202':'E209']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['A214':'C257']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    for row in tips_count_xlsx_ws['E214':'E257']:
        for cell in row:
            try:
                cell.value = None
            except:
                pass

    # 服务组小费
    starting_row = 4
    for team in daily_data[date]["server_group_info"]:
        total_tips = 0
        try:
            total_tips += float(daily_data[date]["server_group_info"][team]["cash_tips"])
        except:
            pass
        try:
            total_tips += float(daily_data[date]["server_group_info"][team]["card_tips"])
        except:
            pass
        try:
            total_tips += float(daily_data[date]["server_group_info"][team]["alipay_tips"])
        except:
            pass
        try:
            total_tips += float(daily_data[date]["server_group_info"][team]["wechatpay_tips"])
        except:
            pass
        if daily_data[date]["server_group_info"][team].get("has_transfer_group") and daily_data[date]["server_group_info"][team]["has_transfer_group"] == "True":
            tips_count_xlsx_ws[f'C{starting_row}'] = total_tips/2
            current_row = starting_row
            for item in daily_data[date]["server_group_info"][team]:
                if "team_member" in item and "work_hour" not in item and daily_data[date]["server_group_info"][team][item] != "":
                    current_team_member = daily_data[date]["server_group_info"][team][item]
                    employee_id = str(current_team_member).split(" ", 1)[0]
                    employee_name = str(current_team_member).split(" ", 1)[1]
                    employee_tips_ratio = employee_data[str(current_team_member).split(" ")[0]]["server_tips_ratio"]
                    employee_work_hour_ratio = float(daily_data[date]["server_group_info"][team][item+"_work_hour"])/8
                    tips_count_xlsx_ws[f'H{current_row}'] = float(employee_work_hour_ratio)
                    tips_count_xlsx_ws[f'A{current_row}'] = employee_id
                    tips_count_xlsx_ws[f'B{current_row}'] = employee_name
                    try:
                        tips_count_xlsx_ws[f'G{current_row}'] = float(employee_tips_ratio)
                    except:
                        tips_count_xlsx_ws[f'G{current_row}'] = 0
                    current_row += 1
            starting_row += 10
            current_row = starting_row
            tips_count_xlsx_ws[f'C{starting_row}'] = total_tips/2
            for item in daily_data[date]["server_group_info"][team]:
                if "transfer_member" in item and "work_hour" not in item and daily_data[date]["server_group_info"][team][item] != "":
                    current_team_member = daily_data[date]["server_group_info"][team][item]
                    employee_id = str(current_team_member).split(" ", 1)[0]
                    employee_name = str(current_team_member).split(" ", 1)[1]
                    employee_tips_ratio = employee_data[str(current_team_member).split(" ")[0]]["server_tips_ratio"]
                    employee_work_hour_ratio = float(daily_data[date]["server_group_info"][team][item+"_work_hour"])/8
                    tips_count_xlsx_ws[f'H{current_row}'] = float(employee_work_hour_ratio)
                    tips_count_xlsx_ws[f'A{current_row}'] = employee_id
                    tips_count_xlsx_ws[f'B{current_row}'] = employee_name
                    try:
                        tips_count_xlsx_ws[f'G{current_row}'] = float(employee_tips_ratio)
                    except:
                        tips_count_xlsx_ws[f'G{current_row}'] = 0
                    current_row += 1
            starting_row += 10
        else:
            tips_count_xlsx_ws[f'C{starting_row}'] = total_tips
            current_row = starting_row
            for item in daily_data[date]["server_group_info"][team]:
                if "team_member" in item and "work_hour" not in item and daily_data[date]["server_group_info"][team][item] != "":
                    current_team_member = daily_data[date]["server_group_info"][team][item]
                    employee_id = str(current_team_member).split(" ", 1)[0]
                    employee_name = str(current_team_member).split(" ", 1)[1]
                    employee_tips_ratio = employee_data[str(current_team_member).split(" ")[0]]["server_tips_ratio"]
                    employee_work_hour_ratio = float(daily_data[date]["server_group_info"][team][item+"_work_hour"])/8
                    tips_count_xlsx_ws[f'H{current_row}'] = float(employee_work_hour_ratio)
                    tips_count_xlsx_ws[f'A{current_row}'] = employee_id
                    tips_count_xlsx_ws[f'B{current_row}'] = employee_name
                    try:
                        tips_count_xlsx_ws[f'G{current_row}'] = float(employee_tips_ratio)
                    except:
                        tips_count_xlsx_ws[f'G{current_row}'] = 0
                    current_row += 1
            starting_row += 10

    # 门迎小费
    starting_row = 108
    for person in daily_data[date]["host_work_hour"]:
        employee_id = str(person).split(" ", 1)[0]
        employee_name = str(person).split(" ", 1)[1]
        employee_tips_ratio = employee_data[str(person).split(" ")[0]]["host_tips_ratios"]
        employee_work_hour = daily_data[date]["host_work_hour"][person]
        tips_count_xlsx_ws[f'A{starting_row}'] = employee_id
        tips_count_xlsx_ws[f'B{starting_row}'] = employee_name
        try:
            tips_count_xlsx_ws[f'C{starting_row}'] = float(employee_tips_ratio)
        except:
            tips_count_xlsx_ws[f'C{starting_row}'] = 0
        tips_count_xlsx_ws[f'E{starting_row}'] = float(employee_work_hour)
        starting_row += 1

    # 保洁小费
    starting_row = 129
    for person in daily_data[date]["cleaner_work_hour"]:
        employee_id = str(person).split(" ", 1)[0]
        employee_name = str(person).split(" ", 1)[1]
        employee_tip_ratio = employee_data[str(person).split(" ")[0]]["cleaner_tips_ratio"]
        employee_work_hour = daily_data[date]["cleaner_work_hour"][person]
        tips_count_xlsx_ws[f'A{starting_row}'] = employee_id
        tips_count_xlsx_ws[f'B{starting_row}'] = employee_name
        try:
            tips_count_xlsx_ws[f'C{starting_row}'] = float(employee_tip_ratio)
        except:
            tips_count_xlsx_ws[f'C{starting_row}'] = 0
        tips_count_xlsx_ws[f'E{starting_row}'] = float(employee_work_hour)
        starting_row += 1

    # 管理及特岗小费
    front_manager = daily_data[date]["manager_daily_work_info_info"]["front_manager"]
    duty_manager = daily_data[date]["manager_daily_work_info_info"]["duty_manager"]
    kitchen_manager = daily_data[date]["manager_daily_work_info_info"]["kitchen_manager"]
    face_changing_master = daily_data[date]["manager_daily_work_info_info"]["face_changing_master"]
    employee_meal_cooker = daily_data[date]["manager_daily_work_info_info"]["employee_meal_cooker"]
    clerk = daily_data[date]["manager_daily_work_info_info"]["clerk"]
    quality_inspector = daily_data[date]["manager_daily_work_info_info"]["quality_inspector"]
    expense_accountant = daily_data[date]["manager_daily_work_info_info"]["expense_accountant"]
    tips_count_xlsx_ws[f'A142'] = str(front_manager).split(" ", 1)[0]
    tips_count_xlsx_ws[f'B142'] = str(front_manager).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A143'] = str(duty_manager).split(" ", 1)[0]
    tips_count_xlsx_ws[f'B143'] = str(duty_manager).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A144'] = str(kitchen_manager).split(" ", 1)[0]
    tips_count_xlsx_ws[f'B144'] = str(kitchen_manager).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A146'] = str(face_changing_master).split(" ", 1)[0]
    tips_count_xlsx_ws[f'B146'] = str(face_changing_master).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A147'] = str(employee_meal_cooker).split(" ", 1)[0]
    tips_count_xlsx_ws[f'B147'] = str(employee_meal_cooker).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A148'] = str(clerk).split(" ", 1)[0]
    tips_count_xlsx_ws[f'B148'] = str(clerk).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A149'] = str(quality_inspector).split(" ", 1)[0]
    tips_count_xlsx_ws[f'B149'] = str(quality_inspector).split(" ", 1)[1]
    tips_count_xlsx_ws[f'A150'] = str(expense_accountant).split(" ", 1)[0]
    tips_count_xlsx_ws[f'B150'] = str(expense_accountant).split(" ", 1)[1]

    # 上菜口小费
    starting_row = 156
    for item in daily_data[date]["dish_prepare_work_hour"]:
        if not "work_hour" in item and daily_data[date]["dish_prepare_work_hour"][item] != "":
            employee_id = str(daily_data[date]["dish_prepare_work_hour"][item]).split(" ", 1)[0]
            employee_name = str(daily_data[date]["dish_prepare_work_hour"][item]).split(" ", 1)[1]
            employee_tip_ratio = employee_data[str(daily_data[date]["dish_prepare_work_hour"][item]).split(" ", 1)[0]]["dish_prepare_tips_ratio"]
            employee_work_hour = daily_data[date]["dish_prepare_work_hour"][item+"_work_hour"]
            tips_count_xlsx_ws[f'A{starting_row}'] = employee_id
            tips_count_xlsx_ws[f'B{starting_row}'] = employee_name
            try:
                tips_count_xlsx_ws[f'C{starting_row}'] = float(employee_tip_ratio)
            except:
                tips_count_xlsx_ws[f'C{starting_row}'] = 0
            tips_count_xlsx_ws[f'E{starting_row}'] = float(employee_work_hour)
            starting_row += 1

    # 小吃房小费
    starting_row = 177
    for person in daily_data[date]["appetizer_work_hour"]:
        employee_id = person.split(" ", 1)[0]
        employee_name = person.split(" ", 1)[1]
        employee_tip_ratio = employee_data[str(person.split(" ", 1)[0])]["appetizer_tips_ratio"]
        employee_work_hour = daily_data[date]["appetizer_work_hour"][person]
        tips_count_xlsx_ws[f'A{starting_row}'] = employee_id
        tips_count_xlsx_ws[f'B{starting_row}'] = employee_name
        try:
            tips_count_xlsx_ws[f'C{starting_row}'] = float(employee_tip_ratio)
        except:
            tips_count_xlsx_ws[f'C{starting_row}'] = 0
        tips_count_xlsx_ws[f'E{starting_row}'] = float(employee_work_hour)
        starting_row += 1

    # 捞面小费
    starting_row = 202
    for person in daily_data[date]["noodle_dance_work_hour"]:
        employee_id = person.split(" ", 1)[0]
        employee_name = person.split(" ", 1)[1]
        employee_tip_ratio = employee_data[str(person.split(" ", 1)[0])]["noodle_dance_tips_ratios"]
        employee_work_hour = daily_data[date]["noodle_dance_work_hour"][person]
        tips_count_xlsx_ws[f'A{starting_row}'] = employee_id
        tips_count_xlsx_ws[f'B{starting_row}'] = employee_name
        try:
            tips_count_xlsx_ws[f'C{starting_row}'] = float(employee_tip_ratio)
        except:
            tips_count_xlsx_ws[f'C{starting_row}'] = 0
        tips_count_xlsx_ws[f'E{starting_row}'] = float(employee_work_hour)
        starting_row += 1

    # 传配小费
    starting_row = 214
    for person in daily_data[date]["dish_runner_work_hour"]:
        employee_id = person.split(" ", 1)[0]
        employee_name = person.split(" ", 1)[1]
        employee_tip_ratio = employee_data[str(person.split(" ", 1)[0])]["dish_runner_tips_ratio"]
        employee_work_hour = float(daily_data[date]["dish_runner_work_hour"][person]["morning_worked_hour"]) + float(daily_data[date]["dish_runner_work_hour"][person]["evening_worked_hour"])
        tips_count_xlsx_ws[f'A{starting_row}'] = employee_id
        tips_count_xlsx_ws[f'B{starting_row}'] = employee_name
        try:
            tips_count_xlsx_ws[f'C{starting_row}'] = float(employee_tip_ratio)
        except:
            tips_count_xlsx_ws[f'C{starting_row}'] = 0
        tips_count_xlsx_ws[f'E{starting_row}'] = float(employee_work_hour)
        starting_row += 1

    tips_count_xlsx_wb.save(tips_count_xlsx_file_path)
    tips_count_xlsx_wb.close()


def get_data_from_daily_data_json_file():
    daily_data = eval(str(requests.get(f"{website_url}/get_daily_data_from_server").json()))
    return daily_data


def get_data_from_employee_data_json_file():
    employee_data = eval(str(requests.get(f"{website_url}/get_employee_data_from_server").json()))
    return employee_data


def init_files():
    print("创建文件夹")
    for file in os.listdir("downloaded_files"):
        os.remove("downloaded_files/" + file)
    try:
        shutil.rmtree(destination_path)
    except:
        pass
    time.sleep(5)
    print("拷贝模板文档")
    shutil.copytree((os.getcwd() + "\\templates"), destination_path)
    for files in os.listdir(destination_path):
        os.rename(destination_path + "\\" + files, destination_path + "\\" + YMD + files)
    print("清除日结表表格内之前的数据")
    copied_close_book_path = None  # 收账本模板
    for file in os.listdir(destination_path + f"\\{YMD}收账明细"):
        if "收账本" in file:
            copied_close_book_path = destination_path + f"\\{YMD}收账明细\\" + file
    copied_close_book_wb = openpyxl.load_workbook(copied_close_book_path)
    copied_daily_close_book_workbook_ws = copied_close_book_wb.active
    for row in copied_daily_close_book_workbook_ws['B1':'B3']:
        for cell in row:
            cell.value = None
    for row in copied_daily_close_book_workbook_ws['C7':'G21']:
        for cell in row:
            cell.value = None
    for row in copied_daily_close_book_workbook_ws['F3':'F4']:
        for cell in row:
            cell.value = None
    for row in copied_daily_close_book_workbook_ws['L3':'O10']:
        for cell in row:
            cell.value = None
    for row in copied_daily_close_book_workbook_ws['R1':'R5']:
        for cell in row:
            cell.value = None
    for row in copied_daily_close_book_workbook_ws["I7":"J21"]:
        for cell in row:
            cell.value = None
    copied_close_book_wb.save(copied_close_book_path)
    copied_close_book_wb.close()


def init_web_driver():
    print("打开网页驱动器")
    s = Service(ChromeDriverManager().install())  # 下载chrome drive
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--lang=en")
    prefs = {'download.default_directory': os.getcwd() + "\\downloaded_files"}
    chrome_options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(chrome_options=chrome_options, executable_path="chromedriver.exe", service=s)
    driver.maximize_window()
    return driver


if __name__ == '__main__':
    date = ""
    while True:
        date = input("请输入日期（yyyymmdd）： ")
        if len(date) != 8 or not date.isdigit():
            print("日期格式有误")
        else:
            break
    YMD = date
    YEAR = date[0:4]
    MONTH = date[4:6]
    DAY = date[6:8]
    NextDay_Date = str(datetime.datetime(int(YEAR), int(MONTH), int(DAY)) + datetime.timedelta(days=1)).split(" ")[0]
    NextDay_year = NextDay_Date.split("-")[0]
    NextDay_month = NextDay_Date.split("-")[1]
    NextDay_day = NextDay_Date.split("-")[2]
    NextDay_Date = NextDay_Date.replace("-", "")
    destination_path = os.getcwd() + "\\daily_files\\" + YEAR + MONTH + DAY
    init_files()

    print("开始从服务器数据导入报表")
    busser_work_count()
    kitchen_work_count()
    daily_tips_count()
    '''
    print("开始海底捞日常清算")
    try:
        pos_haidilao_daily()
    except Exception as e:
        print("!!!海底捞日结失败!!!")
        print(e)

    print("开始IOT日常清算")
    try:
        iot_daily()
    except Exception as e:
        print("!!!IOT日结失败!!!")
        print(e)

    print("开始UberEats日常清算")
    try:
        uber_eat_daily()
    except Exception as e:
        print("!!!uber eat日结失败!!!")
        print(e)

    print("开始Snappy日常清算")
    try:
        snappy_daily()
    except Exception as e:
        print("!!!Snappy日结失败!!!")
        print(e)

    print("开始饭团日常清算")
    try:
        fantuan_daily()
    except Exception as e:
        print("!!!饭团日结失败!!!")
        print(e)
    '''
    input("程序已结束，按enter关闭程序...")
    quit()
